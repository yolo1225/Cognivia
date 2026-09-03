"""Fill a freshly downloaded question template from the submission question source."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = PROJECT_ROOT / "data" / "submission_fixtures" / "ai_app_dev_v1" / "template_question_source.json"
SHEET_NAME = "题目"
HEADERS = [
    "slot_key", "question_external_id", "knowledge_ref", "知识点名称", "purpose", "quiz_level",
    "domain_code", "knowledge_catalog_fingerprint", "题目类型", "难度", "题干", "选项A",
    "选项B", "选项C", "选项D", "正确答案", "解析", "评分点",
]


def read_json(path: Path) -> list[dict[str, Any]]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, list):
        raise ValueError("question source must be a JSON array")
    return value


def answer_value(question: dict[str, Any]) -> str:
    answer_key = question["answer_key"]
    if question["question_type"] == "single_choice":
        return chr(ord("A") + int(answer_key["correct_option"]))
    return str(answer_key.get("answer") or "；".join(answer_key.get("rubric") or []))


def fill_template(template: Path, output: Path, source: Path) -> dict[str, int]:
    if template.resolve() == output.resolve():
        raise ValueError("output must be different from the downloaded template")
    questions = read_json(source)
    source_by_slot = {
        (str(item["knowledge_id"]), str(item["slot_key"])): item
        for item in questions
    }
    if len(source_by_slot) != 450:
        raise ValueError("question source must contain 450 unique knowledge/slot rows")
    workbook = load_workbook(template)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError("question template is missing the 题目 worksheet")
    sheet = workbook[SHEET_NAME]
    headers = [sheet.cell(1, column).value for column in range(1, len(HEADERS) + 1)]
    if headers != HEADERS:
        raise ValueError("question template headers do not match the current import contract")
    columns = {header: index + 1 for index, header in enumerate(HEADERS)}
    consumed: set[tuple[str, str]] = set()
    for row_number in range(2, sheet.max_row + 1):
        knowledge_id = str(sheet.cell(row_number, columns["knowledge_ref"]).value or "").strip()
        slot_key = str(sheet.cell(row_number, columns["slot_key"]).value or "").strip()
        if not knowledge_id and not slot_key:
            continue
        key = (knowledge_id, slot_key)
        question = source_by_slot.get(key)
        if question is None:
            raise ValueError(f"template contains an unexpected slot: {knowledge_id}/{slot_key}")
        if sheet.cell(row_number, columns["purpose"]).value != question["purpose"] or sheet.cell(row_number, columns["quiz_level"]).value != question["quiz_level"]:
            raise ValueError(f"template slot metadata changed: {knowledge_id}/{slot_key}")
        values = {
            "题目类型": question["question_type"],
            "难度": question["difficulty"],
            "题干": question["stem"],
            "选项A": question["options"][0] if question["question_type"] == "single_choice" else "",
            "选项B": question["options"][1] if question["question_type"] == "single_choice" else "",
            "选项C": question["options"][2] if question["question_type"] == "single_choice" else "",
            "选项D": question["options"][3] if question["question_type"] == "single_choice" else "",
            "正确答案": answer_value(question),
            "解析": str(question["answer_key"].get("explanation") or ""),
            "评分点": "\n".join(question["answer_key"].get("rubric") or []),
        }
        for header, value in values.items():
            sheet.cell(row_number, columns[header]).value = value
        consumed.add(key)
    if consumed != set(source_by_slot):
        missing = sorted(set(source_by_slot) - consumed)
        raise ValueError(f"template did not expose all required slots: {missing[:3]}")
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return {"rows_filled": len(consumed), "output": str(output)}


def main() -> None:
    parser = argparse.ArgumentParser(description="Fill a current question-import template from the submission fixture.")
    parser.add_argument("--template", required=True, type=Path, help="XLSX downloaded from the current system")
    parser.add_argument("--output", required=True, type=Path, help="new filled XLSX path")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    args = parser.parse_args()
    result = fill_template(args.template, args.output, args.source)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
