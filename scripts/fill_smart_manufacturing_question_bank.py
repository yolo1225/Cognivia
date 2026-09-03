"""Generate officecli batch chunks for the smart-manufacturing question template."""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl


PROJECT = Path(__file__).resolve().parents[1]
DATASET = PROJECT.parent / "数据集"
TEMPLATE = DATASET / "smart_manufacturing-question-bank-template.xlsx"
KNOWLEDGE_DIR = DATASET / "knowledge"
SOURCE_MD = DATASET / "01-smart-manufacturing-complete.md"
OUT_DIR = PROJECT / "tmp-smart-question-batches"


def load_sources() -> dict[str, dict[str, str]]:
    by_name: dict[str, dict[str, str]] = {}
    for path in KNOWLEDGE_DIR.glob("kb-*.json"):
        for item in json.loads(path.read_text(encoding="utf-8")):
            by_name[str(item["name"])] = {
                "content": str(item.get("content") or ""),
                "source": str(item.get("source_title") or "领域知识包正文"),
            }
    text = SOURCE_MD.read_text(encoding="utf-8")
    sections = re.findall(
        r"^##\s+\d+\.\s+(.+?)\s*$\n(.*?)(?=^##\s+\d+\.\s+|\Z)",
        text,
        flags=re.MULTILINE | re.DOTALL,
    )
    for title, body in sections:
        if title not in by_name:
            by_name[title.strip()] = {
                "content": body,
                "source": "智能制造实训完整知识包",
            }
    return by_name


def facts(content: str) -> list[str]:
    clean = re.sub(r"```.*?```", " ", content, flags=re.DOTALL)
    clean = re.sub(r"^- \*\*.*?$", " ", clean, flags=re.MULTILINE)
    clean = re.sub(r"\s+", " ", clean).strip()
    values: list[str] = []
    for sentence in re.findall(r"[^。！？]+[。！？]", clean):
        sentence = sentence.strip(" -\t")
        if len(sentence) < 16 or sentence in values:
            continue
        values.append(sentence)
        if len(values) == 3:
            break
    if len(values) < 3:
        fallback = clean[:220].rstrip("。！？") + "。"
        while len(values) < 3:
            values.append(fallback)
    return values


def row_content(name: str, slot: str, level: str, content: str) -> dict[str, str | int]:
    f0, f1, f2 = facts(content)
    distractors = [
        "只需把全部样本复制为同一个值",
        "只用于保存训练日志，与该知识点无关",
        "不需要任何输入、配置或现场约束",
    ]
    if slot == "diagnosis_1":
        stem, correct = f"关于{name}，下列说法正确的是？", f0
        qtype, difficulty = "single_choice", 1
    elif slot == "graded_foundation":
        stem, correct = f"学习{name}时，应把握的关键要求是？", f1
        qtype, difficulty = "single_choice", 1
    elif slot == "graded_improvement":
        stem, correct = f"在实践中使用{name}，以下做法最恰当的是？", f2
        qtype, difficulty = "single_choice", 2
    elif slot == "graded_challenge":
        stem, correct = f"若要正确落实{name}，优先应遵循哪项原则？", f"{f0}；并且{f2}"
        qtype, difficulty = "single_choice", 3
    elif slot == "mastery_1":
        stem, correct = f"请概述{name}的核心机制，并说明其主要用途。", f"{f0}{f1}"
        qtype, difficulty = "short_answer", 2
    else:
        stem, correct = (
            f"在一个实际任务中，你会如何应用{name}？请说明理由和关键检查点。",
            f"先依据任务落实：{f1}实施时重点检查：{f2}其依据是：{f0}",
        )
        qtype, difficulty = "short_answer", 4
    if qtype == "single_choice":
        options = [correct, *distractors]
        # Rotate the correct answer to avoid a fixed answer position while
        # keeping the row's answer key explicit.
        shift = {"diagnosis_1": 0, "graded_foundation": 1, "graded_improvement": 2, "graded_challenge": 3}[slot]
        options = options[shift:] + options[:shift]
        answer = "ABCD"[options.index(correct)]
        return {
            "I": qtype, "J": difficulty, "K": stem,
            "L": options[0], "M": options[1], "N": options[2], "O": options[3],
            "P": answer, "Q": f"正确。知识包（{name}）明确指出：{correct}", "R": "",
        }
    rubric = "\n".join([f0, f1, f2])
    return {
        "I": qtype, "J": difficulty, "K": stem,
        "L": "", "M": "", "N": "", "O": "", "P": correct,
        "Q": "参考答案应覆盖核心机制、使用方法和实践要点。", "R": rubric,
    }


def main() -> None:
    sources = load_sources()
    workbook = openpyxl.load_workbook(TEMPLATE, data_only=True)
    sheet = workbook["题目"]
    ops: list[dict[str, object]] = []
    missing: list[str] = []
    for row in range(2, sheet.max_row + 1):
        name = str(sheet.cell(row, 4).value or "")
        source = sources.get(name)
        if source is None:
            missing.append(name)
            continue
        slot = str(sheet.cell(row, 1).value or "")
        values = row_content(name, slot, str(sheet.cell(row, 6).value or ""), source["content"])
        for col, value in values.items():
            ops.append({"command": "set", "path": f"/题目/{col}{row}", "props": {"value": value}})
    if missing:
        raise SystemExit(f"missing_sources:{missing}")
    OUT_DIR.mkdir(exist_ok=True)
    for old in OUT_DIR.glob("*.json"):
        old.unlink()
    for index in range(0, len(ops), 80):
        chunk = ops[index : index + 80]
        (OUT_DIR / f"batch-{index // 80 + 1:03d}.json").write_text(
            json.dumps(chunk, ensure_ascii=False), encoding="utf-8"
        )
    print(json.dumps({"rows": sheet.max_row - 1, "operations": len(ops), "batches": (len(ops) + 79) // 80}, ensure_ascii=False))


if __name__ == "__main__":
    main()
