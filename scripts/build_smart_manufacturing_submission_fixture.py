"""Build the smart-manufacturing database slice from submitted Markdown and XLSX."""

from __future__ import annotations

from collections import Counter, defaultdict
import hashlib
import json
from pathlib import Path
import re
import shutil
from typing import Any

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
FIXTURE_ROOT = PROJECT_ROOT / "data" / "submission_fixtures" / "smart_manufacturing_v1"
SOURCE_MARKDOWN = PROJECT_ROOT / "deliverables" / "knowledge-import-packages" / "smart_manufacturing" / "01-smart-manufacturing-complete.md"
SOURCE_QUESTIONS = PROJECT_ROOT / "deliverables" / "smart_manufacturing-question-bank-filled.xlsx"
FIXTURE_VERSION = "smart_manufacturing_submission_fixture_v1"
PURPOSE_SLOTS = (
    ("diagnosis", "diagnosis_1", "foundation"),
    ("graded_quiz", "graded_foundation", "foundation"),
    ("graded_quiz", "graded_improvement", "improvement"),
    ("graded_quiz", "graded_challenge", "challenge"),
    ("mastery_validation", "mastery_1", "improvement"),
    ("mastery_validation", "mastery_2", "challenge"),
)
ABILITY_DIMENSIONS = ("theory", "practice", "problem_solving", "knowledge_breadth", "learning_speed")


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _metadata_and_content(section: str) -> tuple[dict[str, Any], str]:
    metadata: dict[str, Any] = {}
    content: list[str] = []
    for line in section.splitlines():
        match = re.match(r"^- \*\*([a-z_]+):\*\*\s*(.*?)\s*$", line.strip())
        if match is None:
            content.append(line)
            continue
        key, value = match.groups()
        value = value.strip().strip("`")
        if key in {"tags", "prerequisites"}:
            metadata[key] = [part.strip().strip("`") for part in value.split(",") if part.strip()]
        elif key == "difficulty":
            metadata[key] = int(value)
        elif key == "ability_weights":
            metadata[key] = json.loads(value)
        elif key == "source":
            link = re.match(r"\[([^]]+)]\(([^)]+)\)", value)
            metadata["source_title"] = link.group(1) if link else value
            metadata["source_url"] = link.group(2) if link else None
        elif key == "license":
            metadata["license_note"] = value
        else:
            metadata[key] = value
    return metadata, "\n".join(content).strip()


def _evidence_capabilities(content: str) -> list[str]:
    result = ["concept"]
    if any(marker in content for marker in ("操作步骤", "安装", "配置", "启动", "示教", "排查")):
        result.append("operation")
    if any(marker in content for marker in ("预期结果", "验证", "检查", "可观察")):
        result.append("expected_result")
    if any(marker in content for marker in ("常见错误", "错误", "异常", "故障", "失败")):
        result.append("error_diagnosis")
    return result


def _read_question_workbook() -> tuple[list[dict[str, Any]], dict[str, str], dict[str, str]]:
    workbook = openpyxl.load_workbook(SOURCE_QUESTIONS, read_only=True, data_only=True)
    try:
        rows = workbook["题目"].iter_rows(values_only=True)
        headers = [str(value) for value in next(rows)]
        records = [dict(zip(headers, row, strict=True)) for row in rows]
        metadata = {
            str(key): str(value)
            for key, value in workbook["元数据"].iter_rows(min_row=1, max_col=2, values_only=True)
            if key is not None
        }
    finally:
        workbook.close()
    expected_catalog = "sha256:83969cdc879536b399aa1fef27bb7c3fce32ab8a89ec9d377d532b835b8616f7"
    if metadata.get("domain_code") != "smart_manufacturing" or metadata.get("knowledge_catalog_fingerprint") != expected_catalog:
        raise ValueError("smart manufacturing workbook metadata mismatch")

    by_name: dict[str, str] = {}
    questions: list[dict[str, Any]] = []
    for row in records:
        external_id = str(row["question_external_id"])
        knowledge_id = str(row["knowledge_ref"])
        name = str(row["知识点名称"])
        if by_name.setdefault(name, knowledge_id) != knowledge_id:
            raise ValueError(f"knowledge name maps to multiple IDs: {name}")
        question_type = str(row["题目类型"])
        purpose = str(row["purpose"])
        answer = str(row["正确答案"] or "").strip()
        options = [str(row[column] or "").strip() for column in ("选项A", "选项B", "选项C", "选项D")]
        answer_key: dict[str, Any] = {
            "question_bank_uses": [purpose],
            "quiz_level": str(row["quiz_level"]),
            "explanation": str(row["解析"] or "").strip(),
        }
        if question_type == "single_choice":
            if answer not in {"A", "B", "C", "D"} or any(not option for option in options):
                raise ValueError(f"invalid choice question: {external_id}")
            index = "ABCD".index(answer)
            answer_key.update({"correct_option": index, "answer": options[index]})
        elif question_type == "short_answer":
            rubric = [value.strip() for value in str(row["评分点"] or "").splitlines() if value.strip()]
            if not answer or not 2 <= len(rubric) <= 8:
                raise ValueError(f"invalid short answer question: {external_id}")
            options = []
            answer_key.update({"answer": answer, "rubric": rubric})
        else:
            raise ValueError(f"unsupported question type: {question_type}")
        questions.append(
            {
                "question_id": f"smq_{hashlib.sha256(external_id.encode()).hexdigest()[:20]}",
                "question_external_id": external_id,
                "knowledge_id": knowledge_id,
                "related_knowledge_ids": [],
                "question_type": question_type,
                "difficulty": int(row["难度"]),
                "stem": str(row["题干"]),
                "options": options,
                "answer_key": answer_key,
                "status": "active",
            }
        )
    purpose_counts = Counter(question["answer_key"]["question_bank_uses"][0] for question in questions)
    if len(questions) != 402 or purpose_counts != Counter({"diagnosis": 67, "graded_quiz": 201, "mastery_validation": 134}):
        raise ValueError("question workbook does not contain the expected 402-question inventory")
    return sorted(questions, key=lambda row: row["question_id"]), by_name, metadata


def _read_source_items(knowledge_by_name: dict[str, str]) -> list[dict[str, Any]]:
    text = SOURCE_MARKDOWN.read_text(encoding="utf-8")
    sections = re.findall(r"^## (?!#)(.+?)\s*$\n(.*?)(?=^## (?!#)|\Z)", text, flags=re.MULTILINE | re.DOTALL)
    items: list[dict[str, Any]] = []
    for title, section in sections:
        metadata, content = _metadata_and_content(section)
        name = re.sub(r"^\d+\.\s*", "", title).strip()
        knowledge_id = knowledge_by_name.get(name)
        if knowledge_id is None:
            raise ValueError(f"workbook has no knowledge_ref for: {name}")
        weights = metadata.get("ability_weights")
        if not isinstance(weights, dict) or set(weights) != set(ABILITY_DIMENSIONS):
            raise ValueError(f"invalid ability weights for: {name}")
        if abs(sum(float(weights[key]) for key in ABILITY_DIMENSIONS) - 1.0) > 1e-9:
            raise ValueError(f"ability weights do not sum to one: {name}")
        items.append(
            {
                "knowledge_id": knowledge_id,
                "source_knowledge_id": metadata["knowledge_id"],
                "domain_code": "smart_manufacturing",
                "name": name,
                "category": metadata["category"],
                "difficulty": metadata["difficulty"],
                "tags": metadata["tags"],
                "content": content,
                "source_title": metadata.get("source_title") or "智能制造实训完整知识包",
                "source_url": metadata.get("source_url"),
                "license_note": metadata.get("license_note") or "来源知识包声明",
                "ability_weights": {key: float(weights[key]) for key in ABILITY_DIMENSIONS},
                "evidence_capabilities": _evidence_capabilities(content),
            }
        )
    if len(items) != 67 or len({item["knowledge_id"] for item in items}) != 67:
        raise ValueError("source markdown must resolve exactly 67 unique knowledge items")
    return sorted(items, key=lambda item: item["knowledge_id"])


def _build_relations(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in items:
        grouped[item["category"]].append(item)
    relations: list[dict[str, Any]] = []
    for category, group in sorted(grouped.items()):
        ordered = sorted(group, key=lambda item: (item["difficulty"], item["name"], item["knowledge_id"]))
        for source, target in zip(ordered, ordered[1:], strict=False):
            relations.append(
                {
                    "relation_type": "next_step",
                    "source_knowledge_id": source["knowledge_id"],
                    "target_knowledge_id": target["knowledge_id"],
                    "generation_method": "curriculum_rule",
                    "evidence": {
                        "evidence_kind": "curriculum_rule",
                        "rule": "same_category_difficulty_progression",
                        "category": category,
                        "source_difficulty": source["difficulty"],
                        "target_difficulty": target["difficulty"],
                    },
                }
            )
    return relations


def _build_template_source(questions: list[dict[str, Any]], knowledge_ids: set[str]) -> list[dict[str, Any]]:
    grouped = {
        (
            row["knowledge_id"],
            row["answer_key"]["question_bank_uses"][0],
            row["answer_key"]["quiz_level"],
        ): row
        for row in questions
    }
    template: list[dict[str, Any]] = []
    for knowledge_id in sorted(knowledge_ids):
        for purpose, slot_key, quiz_level in PURPOSE_SLOTS:
            question = grouped.get((knowledge_id, purpose, quiz_level))
            if question is None:
                raise ValueError(f"missing template question: {knowledge_id}/{purpose}")
            template.append({"slot_key": slot_key, "knowledge_id": knowledge_id, "purpose": purpose, "quiz_level": quiz_level, "source_question_id": question["question_id"]})
    if len({row["source_question_id"] for row in template}) != len(template):
        raise ValueError("template source must retain one distinct question per official slot")
    return template


def _build_import_markdown(items: list[dict[str, Any]]) -> str:
    lines = ["# 智能制造实训完整知识包 (smart_manufacturing)", "", "本文件用于智能制造提交夹具的空库导入演示，请与启动夹具二选一使用。", ""]
    for item in items:
        source_title = item["source_title"].replace("]", "\\]")
        source = f"[{source_title}]({item['source_url']})" if item["source_url"] else source_title
        lines.extend([
            f"## {item['name']}",
            f"- **knowledge_id:** `{item['knowledge_id']}`",
            f"- **category:** {item['category']}",
            f"- **difficulty:** {item['difficulty']}",
            f"- **tags:** {', '.join(item['tags'])}",
            f"- **source:** {source}",
            f"- **license:** {item['license_note']}",
            f"- **ability_weights:** `{json.dumps(item['ability_weights'], ensure_ascii=False, separators=(',', ':'))}`",
            "",
            item["content"],
            "",
        ])
    return "\n".join(lines)


def _weak_item(item: dict[str, Any], level: int) -> dict[str, Any]:
    return {
        "knowledge_id": item["knowledge_id"],
        "name": item["name"],
        "category": item["category"],
        "weakness_level": level,
        "mastery_type": "unmastered" if level >= 4 else "partial_mastery",
        "prerequisites": [],
        "evidence_ids": [f"submission_fixture:{item['knowledge_id']}"],
        "reason": "智能制造提交夹具中的脱敏诊断基线",
    }


def _build_learner_profiles(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_name = {item["name"]: item for item in items}
    selected = {
        "beginner": ["OpenPLC Runtime v4 容器化部署与编辑器连接", "PLC 定义与循环扫描工作原理", "工业互联网网络体系"],
        "intermediate": ["项目创建与硬件组态", "机器人与外设协同"],
        "advanced": ["UR ROS 2 Driver 启动与控制模式衔接", "UR External Control 中断识别与恢复"],
    }
    missing = sorted(name for names in selected.values() for name in names if name not in by_name)
    if missing:
        raise ValueError(f"profile target knowledge missing: {missing}")
    version = FIXTURE_VERSION
    return [
        {
            "learner_id": "sm_fixture_learner_beginner", "profile_id": "sm_fixture_profile_beginner_v1", "path_id": "sm_fixture_path_beginner_v1", "domain_code": "smart_manufacturing",
            "education_level": "高职在读", "major": "机电一体化技术", "experience_years": 0, "learning_style": "guided", "direction_tags": ["plc", "industrial_internet"], "background": "高职在读｜机电一体化技术｜0年相关经验",
            "ability_profile": {"profile_type": "beginner", "theory": 38, "practice": 28, "problem_solving": 32, "breadth": 35, "learning_speed": 46, "category_mastery": {}, "blind_spot_ids": []},
            "weak_knowledge": [_weak_item(by_name[name], 5) for name in selected["beginner"]], "confirmed_at": "2026-09-02T00:00:00+00:00", "fixture_version": version,
        },
        {
            "learner_id": "sm_fixture_learner_intermediate", "profile_id": "sm_fixture_profile_intermediate_v1", "path_id": "sm_fixture_path_intermediate_v1", "domain_code": "smart_manufacturing",
            "education_level": "本科", "major": "自动化", "experience_years": 1, "learning_style": "mixed", "direction_tags": ["tia_portal", "robot_integration"], "background": "本科｜自动化｜1年相关经验",
            "ability_profile": {"profile_type": "intermediate", "theory": 66, "practice": 58, "problem_solving": 61, "breadth": 60, "learning_speed": 64, "category_mastery": {}, "blind_spot_ids": []},
            "weak_knowledge": [_weak_item(by_name[name], 4) for name in selected["intermediate"]], "confirmed_at": "2026-09-02T00:00:00+00:00", "fixture_version": version,
        },
        {
            "learner_id": "sm_fixture_learner_advanced", "profile_id": "sm_fixture_profile_advanced_v1", "path_id": "sm_fixture_path_advanced_v1", "domain_code": "smart_manufacturing",
            "education_level": "硕士", "major": "控制工程", "experience_years": 3, "learning_style": "practice_oriented", "direction_tags": ["ros2", "robot_integration"], "background": "硕士｜控制工程｜3年相关经验",
            "ability_profile": {"profile_type": "advanced", "theory": 88, "practice": 84, "problem_solving": 90, "breadth": 82, "learning_speed": 80, "category_mastery": {}, "blind_spot_ids": []},
            "weak_knowledge": [_weak_item(by_name[name], 3) for name in selected["advanced"]], "confirmed_at": "2026-09-02T00:00:00+00:00", "fixture_version": version,
        },
    ]


def _build_manual_demo_cases(profiles: list[dict[str, Any]]) -> dict[str, Any]:
    """Keep the three live-demo inputs versioned without adding offline evaluations."""
    by_type = {item["ability_profile"]["profile_type"]: item for item in profiles}
    cases = []
    for profile_type, case_id, scenario, goal, follow_up in (
        (
            "beginner",
            "SM-DEMO-BEGINNER-INITIAL",
            "PLC 与工业互联网基础薄弱学习者的补救型初始生成",
            "围绕 OpenPLC 仿真部署、PLC 循环扫描与工业互联网网络体系生成可追溯的基础学习包。",
            {"type": "initial_generation"},
        ),
        (
            "intermediate",
            "SM-DEMO-INTERMEDIATE-REVIEW",
            "TIA 组态与机器人 I/O 待巩固学习者提交疑似错误反馈并触发复核",
            "围绕项目创建硬件组态与机器人外设协同生成可追溯的进阶学习包。",
            {"type": "incorrect_feedback", "expected_action": "review", "profile_update_required": False},
        ),
        (
            "advanced",
            "SM-DEMO-ADVANCED-CHALLENGE",
            "UR ROS 2 集成与恢复待提升学习者以已确认掌握证据请求挑战任务",
            "围绕 UR ROS 2 Driver 控制模式与 External Control 恢复生成迁移挑战学习包。",
            {"type": "challenge_request", "expected_action": "challenge"},
        ),
    ):
        profile = by_type[profile_type]
        cases.append(
            {
                "case_id": case_id,
                "scenario": scenario,
                "learner_id": profile["learner_id"],
                "profile_id": profile["profile_id"],
                "profile_type": profile_type,
                "weak_knowledge_ids": [item["knowledge_id"] for item in profile["weak_knowledge"]],
                "learning_goal": goal,
                "resource_types": ["lecture", "practice_guide", "graded_quiz"],
                "follow_up": follow_up,
            }
        )
    return {
        "schema_version": "smart-manufacturing-live-demo-cases-v1",
        "domain_code": "smart_manufacturing",
        "privacy": "all identifiers are synthetic; no full answer text, raw Agent payload, or resource body is retained",
        "cases": cases,
    }


def build() -> None:
    questions, knowledge_by_name, workbook_metadata = _read_question_workbook()
    items = _read_source_items(knowledge_by_name)
    knowledge_ids = {item["knowledge_id"] for item in items}
    if knowledge_ids != {question["knowledge_id"] for question in questions}:
        raise ValueError("knowledge package and question workbook scopes differ")
    relations = _build_relations(items)
    template = _build_template_source(questions, knowledge_ids)
    profiles = _build_learner_profiles(items)
    manual_cases = _build_manual_demo_cases(profiles)
    domain = {
        "domain_code": "smart_manufacturing", "name": "智能制造实训", "domain_schema_version": "1.0",
        "resource_types": ["lecture", "practice_guide", "graded_quiz"], "ability_dimensions": list(ABILITY_DIMENSIONS),
        "learning_directions": [
            {"value": "plc_control", "label": "PLC 控制与组态", "description": "PLC、梯形图和 TIA Portal 实操", "match_tags": ["plc", "tia", "ladder"]},
            {"value": "industrial_connectivity", "label": "工业互联网连接", "description": "网络、标识与工业互联网基础", "match_tags": ["industrial-internet", "network", "iot"]},
            {"value": "robot_integration", "label": "机器人集成", "description": "示教、I/O、ROS 2 与机器人协同", "match_tags": ["robot", "ros2", "driver"]},
        ],
        "mvp_targets": {"knowledge_items": 67, "diagnostic_questions": 402, "evaluation_cases": 0},
        "profile_policy": {"version": "smart_manufacturing_profile_v1", "data_fingerprint": workbook_metadata["knowledge_catalog_fingerprint"]},
    }
    source_dir = FIXTURE_ROOT / "source_assets"
    import_dir = FIXTURE_ROOT / "import_source"
    source_dir.mkdir(parents=True, exist_ok=True)
    import_dir.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(SOURCE_MARKDOWN, source_dir / SOURCE_MARKDOWN.name)
    shutil.copyfile(SOURCE_QUESTIONS, source_dir / SOURCE_QUESTIONS.name)
    import_path = import_dir / "01-smart-manufacturing-complete.md"
    import_path.write_text(_build_import_markdown(items), encoding="utf-8")
    write_json(FIXTURE_ROOT / "domain.json", domain)
    write_json(FIXTURE_ROOT / "knowledge_items.json", items)
    write_json(FIXTURE_ROOT / "relations.json", relations)
    write_json(FIXTURE_ROOT / "diagnostic_questions.json", questions)
    write_json(FIXTURE_ROOT / "template_question_source.json", template)
    write_json(FIXTURE_ROOT / "supplemental_diagnosis_questions.json", [])
    write_json(FIXTURE_ROOT / "learner_profiles.json", profiles)
    write_json(FIXTURE_ROOT / "manual_demo_cases.json", manual_cases)
    write_json(FIXTURE_ROOT / "import_source_manifest.json", {
        "schema_version": "smart-manufacturing-import-source-v1", "domain_code": "smart_manufacturing", "mode": "empty_database_import_demo", "knowledge_count": 67, "file": import_path.name, "sha256": sha256(import_path), "source_fixture_version": FIXTURE_VERSION,
        "note": "Use this canonical import source or the bootstrap fixture, never both in one database.",
    })
    (FIXTURE_ROOT / "README.md").write_text(
        "# 智能制造可执行测试数据包\n\n本目录用于从空库复现比赛第二领域 `smart_manufacturing`。它包含数据库切片、三类脱敏学情以及受管的三案例运行输入；不包含 50 例离线评测案例，也不声明第二领域质量指标。\n\n## 使用方式\n\n1. 运行 `scripts/submission-fixture.ps1 verify -FixtureDir data/submission_fixtures/smart_manufacturing_v1` 校验哈希和内容。\n2. 在新的 Docker 卷或已清空数据库中运行 `scripts/submission-fixture.ps1 bootstrap -FixtureDir data/submission_fixtures/smart_manufacturing_v1`。脚本不会清空现有数据。\n3. 若需与主演示环境隔离，可添加 `-ComposeProject cognivia_sm_test -ComposeFile docker-compose.submission.yml`。\n4. 构建索引后运行 `python test_script/smart_manufacturing_demo_acceptance.py --base-url http://localhost:18000/api/v1`，生成脱敏案例与报告。\n5. `import_source/` 中的 Markdown 与启动夹具互斥，不能在同一数据库叠加导入。\n\n## 内容\n\n- 67 条可追溯知识条目与课程规则生成的 `next_step` 图谱关系。\n- 402 道活动正式题：67 道诊断题、201 道分阶测验题、134 道掌握检查题。\n- 初学者、中阶和高阶三份合成画像及学习路径；所有账号均标记为测试数据。\n- `manual_demo_cases.json`：三组真实运行的受管输入与预期业务断言。\n- `source_assets/` 保存受管 Markdown 和 XLSX 的哈希锁定副本。\n",
        encoding="utf-8",
    )
    files = (
        "domain.json", "knowledge_items.json", "relations.json", "diagnostic_questions.json", "template_question_source.json", "supplemental_diagnosis_questions.json", "learner_profiles.json", "manual_demo_cases.json", "import_source_manifest.json", "import_source/01-smart-manufacturing-complete.md", "source_assets/01-smart-manufacturing-complete.md", "source_assets/smart_manufacturing-question-bank-filled.xlsx", "README.md",
    )
    relation_counts = Counter(relation["relation_type"] for relation in relations)
    purpose_counts = Counter(question["answer_key"]["question_bank_uses"][0] for question in questions)
    manifest = {
        "schema_version": "submission-fixture-manifest-v1", "fixture_version": FIXTURE_VERSION, "domain_code": "smart_manufacturing",
        "source_assets": {"knowledge_package": "source_assets/01-smart-manufacturing-complete.md", "question_workbook": "source_assets/smart_manufacturing-question-bank-filled.xlsx", "knowledge_catalog_fingerprint": workbook_metadata["knowledge_catalog_fingerprint"], "question_inventory_fingerprint": workbook_metadata["question_inventory_fingerprint"]},
        "counts": {"knowledge_items": len(items), "knowledge_relations": len(relations), "relation_types": dict(sorted(relation_counts.items())), "active_questions": len(questions), "question_purposes": dict(sorted(purpose_counts.items())), "template_compatible_questions": len(template), "supplemental_diagnosis_questions": 0, "evaluation_cases": 0, "manual_demo_cases": len(manual_cases["cases"]), "learner_profiles": len(profiles)},
        "files": {name: {"sha256": sha256(FIXTURE_ROOT / name)} for name in files},
        "import_source": {"path": import_path.relative_to(FIXTURE_ROOT).as_posix(), "sha256": sha256(import_path), "deliverable_path": SOURCE_MARKDOWN.relative_to(PROJECT_ROOT).as_posix()},
        "reproduction_modes": {"bootstrap": "load the fixture into an empty database", "import_demo": "upload the canonical Markdown, then import matching formal questions"},
    }
    write_json(FIXTURE_ROOT / "manifest.json", manifest)
    print(json.dumps({"fixture_root": str(FIXTURE_ROOT), "counts": manifest["counts"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    build()
