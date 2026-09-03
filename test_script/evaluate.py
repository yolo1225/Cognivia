from __future__ import annotations

import argparse
import copy
import json
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
CASES_DIR = ROOT / "data" / "evaluation_cases"
CASES_MANIFEST = CASES_DIR / "manifest.json"
REPORT_DIR = ROOT / "reports" / "evaluation"
SCRIPT_VERSION = "live-evaluator-3.0-v6"
END_TO_END_LATENCY_SLA_MS = 120_000
EXPECTED_SCENARIOS = {
    "initial_generation": 40,
    "feedback_revision": 5,
    "challenge_task": 5,
}
EXPECTED_PROFILES = {"beginner", "intermediate", "advanced"}
EXPECTED_RESOURCE_TYPES = {"lecture", "practice_guide", "graded_quiz"}


def load_cases() -> tuple[list[dict[str, Any]], set[str]]:
    manifest = json.loads(CASES_MANIFEST.read_text(encoding="utf-8"))
    active_file = manifest.get("active_file")
    if not isinstance(active_file, str):
        raise ValueError("evaluation case manifest requires active_file")
    path = CASES_DIR / active_file
    payload = json.loads(path.read_text(encoding="utf-8"))
    cases = payload.get("cases", []) if isinstance(payload, dict) else payload
    if not isinstance(cases, list):
        raise ValueError(f"{path.name}: cases must be an array")
    versions = {str(payload["knowledge_base_version"])} if payload.get("knowledge_base_version") else set()
    case_ids = [str(item.get("case_id")) for item in cases]
    if len(case_ids) != len(set(case_ids)):
        raise ValueError("duplicate case_id found")
    if len(cases) != 50:
        raise ValueError(f"V4 evaluation requires exactly 50 cases, found {len(cases)}")
    scenarios = Counter(str(item.get("scenario_type")) for item in cases)
    if scenarios != EXPECTED_SCENARIOS:
        raise ValueError(f"V4 scenario distribution mismatch: {dict(scenarios)}")
    profiles = {
        str((item.get("profile_snapshot") or {}).get("profile_type")) for item in cases
    }
    if profiles != EXPECTED_PROFILES:
        raise ValueError(f"V4 learner profiles mismatch: {sorted(profiles)}")
    resource_types = {str(item.get("resource_type")) for item in cases}
    if resource_types != EXPECTED_RESOURCE_TYPES:
        raise ValueError(f"V4 resource types mismatch: {sorted(resource_types)}")
    return cases, versions


def _ratio(numerator: int, denominator: int) -> dict[str, Any]:
    return {
        "numerator": numerator,
        "denominator": denominator,
        "ratio": round(numerator / denominator, 4) if denominator else None,
    }


def _percentile(values: list[int], percentile: float) -> int | None:
    if not values:
        return None
    values = sorted(values)
    index = max(0, min(len(values) - 1, int((len(values) - 1) * percentile + 0.9999)))
    return values[index]


def _latency_summary(values: list[int]) -> dict[str, Any]:
    return {
        "sample_count": len(values),
        "p50": _percentile(values, 0.50),
        "p95": _percentile(values, 0.95),
    }


def load_live_run(run_id: str | None) -> dict[str, Any]:
    run_dir = REPORT_DIR / "runs"
    if run_id:
        path = run_dir / f"{run_id}.json"
    else:
        candidates = sorted(run_dir.glob("*.json"), key=lambda item: item.stat().st_mtime)
        if not candidates:
            raise FileNotFoundError("no live evaluation run found")
        path = candidates[-1]
    if not path.is_file():
        raise FileNotFoundError(f"live evaluation run not found: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def merge_live_results(
    cases: list[dict[str, Any]], run: dict[str, Any]
) -> list[dict[str, Any]]:
    by_case = {str(item["case_id"]): item for item in run.get("results", [])}
    merged: list[dict[str, Any]] = []
    for source in cases:
        case_id = str(source.get("case_id"))
        if case_id not in by_case:
            continue
        item = copy.deepcopy(source)
        item["observed_result"] = by_case[case_id].get("observed_result", {})
        # Task terminal state is a live execution fact, not part of the gold case.
        # Keep it when merging so the success-rate denominator remains auditable.
        item["task_status"] = by_case[case_id].get("task_status")
        merged.append(item)
    return merged


def evaluate(
    cases: list[dict[str, Any]],
    knowledge_versions: set[str],
    *,
    run_mode: str = "baseline",
    run_metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    run_metadata = run_metadata or {}
    determinable = [item for item in cases if item.get("observed_result", {}).get("determinable")]
    undetermined = [str(item.get("case_id")) for item in cases if item not in determinable]
    facts = sum(int(item["observed_result"].get("generated_fact_count", 0)) for item in determinable)
    hallucinated = sum(
        int(item["observed_result"].get("hallucinated_fact_count", 0)) for item in determinable
    )
    evidence_insufficient = sum(
        int(item["observed_result"].get("evidence_insufficient_claim_count", 0))
        for item in determinable
    )
    unresolved = sum(
        int(item["observed_result"].get("unresolved_claim_count", 0))
        for item in determinable
    )
    difficulty_evaluable = [
        item
        for item in determinable
        if item["observed_result"].get("difficulty_matched") is not None
    ]
    difficulty_pass = [
        str(item["case_id"])
        for item in difficulty_evaluable
        if bool(item["observed_result"].get("difficulty_matched"))
    ]
    coverage_numerator = sum(
        int(item["observed_result"].get("covered_core_knowledge_count", 0))
        for item in determinable
    )
    coverage_denominator = sum(
        int(item["observed_result"].get("target_core_knowledge_count", 0))
        for item in determinable
    )
    review_pass = [
        str(item["case_id"])
        for item in determinable
        if item["observed_result"].get("review_conclusion")
        == item.get("expected_review_conclusion")
    ]
    profile_pass = [
        str(item["case_id"])
        for item in determinable
        if item["observed_result"].get("profile_decision")
        == item.get("expected_profile_decision")
    ]
    latencies = [
        int(item["observed_result"].get("latency_ms", 0))
        for item in determinable
        if item["observed_result"].get("latency_ms") is not None
    ]
    trigger_response_times = [
        int(item["observed_result"].get("trigger_response_ms", 0))
        for item in determinable
        if item["observed_result"].get("trigger_response_ms") is not None
    ]
    task_status_observed = [
        item for item in determinable if item.get("task_status") is not None
    ]
    completed_tasks = [
        item
        for item in task_status_observed
        if item.get("task_status") == "completed"
    ]
    delayed_cases = [
        item
        for item in determinable
        if int(item["observed_result"].get("latency_ms", 0)) > END_TO_END_LATENCY_SLA_MS
    ]
    agent_latency_values: dict[str, list[int]] = {}
    for item in determinable:
        for agent_name, duration in (
            item["observed_result"].get("agent_latency_ms") or {}
        ).items():
            agent_latency_values.setdefault(agent_name, []).append(int(duration))
    difficulty = _ratio(len(difficulty_pass), len(difficulty_evaluable))
    coverage = _ratio(coverage_numerator, coverage_denominator)
    hallucination = _ratio(hallucinated, facts)
    competition_checks = {
        "at_least_50_cases": len(cases) >= 50,
        "all_cases_determinable": len(determinable) == len(cases),
        "hallucination_rate_below_5_percent": (
            hallucination["ratio"] is not None and hallucination["ratio"] < 0.05
        ),
        "difficulty_match_at_least_85_percent": (
            difficulty["ratio"] is not None and difficulty["ratio"] >= 0.85
        ),
        "core_coverage_at_least_90_percent": (
            coverage["ratio"] is not None and coverage["ratio"] >= 0.90
        ),
    }
    diagnostic_checks = {
        "no_evidence_insufficient": evidence_insufficient == 0,
        "no_unresolved_claims": unresolved == 0,
        "review_decisions_match": len(review_pass) == len(determinable),
        "profile_decisions_match": len(profile_pass) == len(determinable),
    }
    result = {
        "status": "passed" if all(competition_checks.values()) else "failed",
        "case_count": len(cases),
        "mvp_target_case_count": 50,
        "evaluated_case_count": len(determinable),
        "competition_acceptance": {
            "accepted": all(competition_checks.values()),
            "competition_checks": competition_checks,
            "diagnostic_checks": diagnostic_checks,
            "failed_checks": [
                name for name, passed in competition_checks.items() if not passed
            ],
            "diagnostic_findings": [
                name for name, passed in diagnostic_checks.items() if not passed
            ],
        },
        "metrics": {
            "hallucination_rate": hallucination,
            "evidence_insufficient_claims": {
                "count": evidence_insufficient,
                "publication_blocking": True,
            },
            "unresolved_claims": {"count": unresolved, "publication_blocking": True},
            "difficulty_match_accuracy": difficulty,
            "difficulty_not_applicable_case_count": len(determinable)
            - len(difficulty_evaluable),
            "core_knowledge_coverage": coverage,
            "review_decision_accuracy": _ratio(len(review_pass), len(determinable)),
            "profile_decision_accuracy": _ratio(len(profile_pass), len(determinable)),
            # `latency_ms` is the end-to-end business workflow duration from the
            # evaluation trigger to a terminal task state. It deliberately
            # includes the review gate, rather than reporting only async enqueue.
            "latency_ms": _latency_summary(latencies),
            "trigger_response_ms": _latency_summary(trigger_response_times),
            "end_to_end_latency_sla": {
                "threshold_ms": END_TO_END_LATENCY_SLA_MS,
                **_ratio(len(delayed_cases), len(determinable)),
            },
            "task_success_rate": _ratio(len(completed_tasks), len(task_status_observed)),
            "agent_latency_ms": {
                name: {"p50": _percentile(values, 0.50), "p95": _percentile(values, 0.95)}
                for name, values in sorted(agent_latency_values.items())
            },
        },
        "failed_case_ids": {
            "hallucination": [
                str(item["case_id"])
                for item in determinable
                if int(item["observed_result"].get("hallucinated_fact_count", 0)) > 0
            ],
            "evidence_insufficient": [
                str(item["case_id"])
                for item in determinable
                if int(item["observed_result"].get("evidence_insufficient_claim_count", 0)) > 0
            ],
            "unresolved": [
                str(item["case_id"])
                for item in determinable
                if int(item["observed_result"].get("unresolved_claim_count", 0)) > 0
            ],
            "difficulty": [
                str(item["case_id"])
                for item in difficulty_evaluable
                if str(item["case_id"]) not in difficulty_pass
            ],
            "coverage": [
                str(item["case_id"])
                for item in determinable
                if int(item["observed_result"].get("covered_core_knowledge_count", 0))
                < int(item["observed_result"].get("target_core_knowledge_count", 0))
            ],
            "review_decision": [str(item["case_id"]) for item in determinable if str(item["case_id"]) not in review_pass],
            "profile_decision": [str(item["case_id"]) for item in determinable if str(item["case_id"]) not in profile_pass],
            "end_to_end_latency_sla": [
                str(item["case_id"]) for item in delayed_cases
            ],
        },
        "unable_to_determine": {
            "count": len(undetermined),
            "case_ids": undetermined,
            "statement": "Cases without a determinable observed result are excluded from metric denominators.",
        },
        "case_results": [
            {
                "case_id": str(item.get("case_id")),
                "scenario_type": item.get("scenario_type"),
                "profile_type": (item.get("profile_snapshot") or {}).get("profile_type"),
                "resource_type": item.get("resource_type"),
                "determinable": bool((item.get("observed_result") or {}).get("determinable")),
                "failure_category": (item.get("observed_result") or {}).get(
                    "failure_category"
                ),
                "failure_code": (item.get("observed_result") or {}).get("failure_code"),
                "classification_basis": (item.get("observed_result") or {}).get(
                    "classification_basis"
                ),
                "field_paths": (item.get("observed_result") or {}).get("field_paths", []),
                "latency_ms": (item.get("observed_result") or {}).get("latency_ms"),
                "trigger_response_ms": (item.get("observed_result") or {}).get(
                    "trigger_response_ms"
                ),
            }
            for item in cases
        ],
        "knowledge_base_versions": sorted(knowledge_versions),
        "run_mode": run_mode,
        "run_id": run_metadata.get("run_id"),
        "stage": run_metadata.get("stage"),
        "diagnostic_case_id": run_metadata.get("diagnostic_case_id"),
        "model_configuration": run_metadata.get("model_configuration", {}),
        "rag_configuration": run_metadata.get("rag_configuration", {}),
        "case_set_sha256": run_metadata.get("case_set_sha256"),
        "full_suite_case_sha256": run_metadata.get("full_suite_case_sha256"),
        "run_complete": run_metadata.get("complete"),
        "run_valid": run_metadata.get("valid"),
        "script_version": SCRIPT_VERSION,
        "evaluated_at": datetime.now(UTC).isoformat(),
    }
    return result


def write_reports(result: dict[str, Any], *, xlsx: bool) -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    mode = result.get("run_mode", "baseline")
    stem = "latest-live" if mode == "live" else "latest-baseline"
    (REPORT_DIR / f"{stem}.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    if mode == "baseline":
        (REPORT_DIR / "latest.json").write_text(
            json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    metrics = result["metrics"]
    lines = [
        f"# {mode.upper()} 可复现评测报告",
        "",
        f"- 状态：{result['status']}",
        f"- 案例：{result['evaluated_case_count']}/{result['case_count']}",
        f"- 知识库版本：{', '.join(result['knowledge_base_versions'])}",
        f"- 脚本版本：{result['script_version']}",
        f"- 评测时间：{result['evaluated_at']}",
        f"- 运行模式：{mode}",
        f"- 运行编号：{result.get('run_id') or 'baseline'}",
        "",
        "| 指标 | 分子 | 分母 | 比率 |",
        "|---|---:|---:|---:|",
    ]
    for label, key in (
        ("幻觉率", "hallucination_rate"),
        ("难度匹配准确率", "difficulty_match_accuracy"),
        ("核心知识覆盖率", "core_knowledge_coverage"),
        ("审核结论准确率", "review_decision_accuracy"),
        ("画像结论准确率", "profile_decision_accuracy"),
    ):
        item = metrics[key]
        lines.append(f"| {label} | {item['numerator']} | {item['denominator']} | {item['ratio']} |")
    lines.extend(
        [
            "",
            "## 性能",
            "",
            "| 指标 | 样本量 | 结果 |",
            "|---|---:|---:|",
            f"| 端到端业务时延 P50 | {metrics['latency_ms']['sample_count']} | {metrics['latency_ms']['p50']} ms |",
            f"| 端到端业务时延 P95 | {metrics['latency_ms']['sample_count']} | {metrics['latency_ms']['p95']} ms |",
            f"| 触发接口确认时延 P50 | {metrics['trigger_response_ms']['sample_count']} | {metrics['trigger_response_ms']['p50'] if metrics['trigger_response_ms']['p50'] is not None else '未采集'} ms |",
            f"| 触发接口确认时延 P95 | {metrics['trigger_response_ms']['sample_count']} | {metrics['trigger_response_ms']['p95'] if metrics['trigger_response_ms']['p95'] is not None else '未采集'} ms |",
            f"| 超 {metrics['end_to_end_latency_sla']['threshold_ms']} ms 延迟率 | {metrics['end_to_end_latency_sla']['denominator']} | {metrics['end_to_end_latency_sla']['numerator']} / {metrics['end_to_end_latency_sla']['denominator']} = {metrics['end_to_end_latency_sla']['ratio']} |",
            f"| 任务成功率 | {metrics['task_success_rate']['denominator']} | {metrics['task_success_rate']['numerator']} / {metrics['task_success_rate']['denominator']} = {metrics['task_success_rate']['ratio']} |",
            "",
            "## 失败案例",
            "",
            *[
                f"- {name}: {', '.join(ids) if ids else '无'}"
                for name, ids in result["failed_case_ids"].items()
            ],
            "",
            f"无法判定：{result['unable_to_determine']['statement']}",
        ]
    )
    classified_failures = [
        item for item in result.get("case_results", []) if item.get("failure_category")
    ]
    if classified_failures:
        lines.extend(
            [
                "",
                "## 失败归因",
                "",
                "| 案例 | 主因 | 终态 | 判定依据 |",
                "|---|---|---|---|",
            ]
        )
        for item in classified_failures:
            basis = str(item.get("classification_basis") or "").replace("|", "\\|")
            lines.append(
                f"| {item['case_id']} | {item['failure_category']} | "
                f"{item.get('failure_code') or ''} | {basis} |"
            )
    agent_latency = metrics.get("agent_latency_ms") or {}
    if agent_latency:
        lines.extend(["", "## Agent 性能", "", "| Agent | P50 ms | P95 ms |", "|---|---:|---:|"])
        for name, values in agent_latency.items():
            lines.append(f"| {name} | {values['p50']} | {values['p95']} |")
    (REPORT_DIR / f"{stem}.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    if mode == "baseline":
        (REPORT_DIR / "latest.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    if xlsx:
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "summary"
        sheet.append(["metric", "numerator", "denominator", "ratio"])
        for key, item in metrics.items():
            if isinstance(item, dict) and "ratio" in item:
                sheet.append([key, item["numerator"], item["denominator"], item["ratio"]])
        sheet.append(["latency_p50_ms", metrics["latency_ms"]["p50"], None, None])
        sheet.append(["latency_p95_ms", metrics["latency_ms"]["p95"], None, None])
        sheet.append(["trigger_response_p50_ms", metrics["trigger_response_ms"]["p50"], None, None])
        sheet.append(["trigger_response_p95_ms", metrics["trigger_response_ms"]["p95"], None, None])
        sheet.append([
            "end_to_end_latency_sla_breach_rate",
            metrics["end_to_end_latency_sla"]["numerator"],
            metrics["end_to_end_latency_sla"]["denominator"],
            metrics["end_to_end_latency_sla"]["ratio"],
        ])
        sheet.append([
            "task_success_rate",
            metrics["task_success_rate"]["numerator"],
            metrics["task_success_rate"]["denominator"],
            metrics["task_success_rate"]["ratio"],
        ])
        agent_sheet = workbook.create_sheet("agent_latency")
        agent_sheet.append(["agent", "p50_ms", "p95_ms"])
        for name, values in metrics.get("agent_latency_ms", {}).items():
            agent_sheet.append([name, values["p50"], values["p95"]])
        case_sheet = workbook.create_sheet("case_results")
        case_sheet.append(
            [
                "case_id",
                "scenario_type",
                "profile_type",
                "resource_type",
                "determinable",
                "failure_category",
                "failure_code",
                "classification_basis",
                "field_paths",
                "latency_ms",
                "trigger_response_ms",
            ]
        )
        for item in result.get("case_results", []):
            case_sheet.append(
                [
                    item.get("case_id"),
                    item.get("scenario_type"),
                    item.get("profile_type"),
                    item.get("resource_type"),
                    item.get("determinable"),
                    item.get("failure_category"),
                    item.get("failure_code"),
                    item.get("classification_basis"),
                    json.dumps(item.get("field_paths") or [], ensure_ascii=False),
                    item.get("latency_ms"),
                    item.get("trigger_response_ms"),
                ]
            )
        workbook.save(REPORT_DIR / f"{stem}.xlsx")
        if mode == "baseline":
            workbook.save(REPORT_DIR / "latest.xlsx")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", action="store_true", help="also export latest.xlsx")
    parser.add_argument("--mode", choices=("baseline", "live"), default="baseline")
    parser.add_argument("--run-id", help="live run id; latest run is used when omitted")
    args = parser.parse_args()
    cases, versions = load_cases()
    run: dict[str, Any] = {}
    if args.mode == "live":
        run = load_live_run(args.run_id)
        cases = merge_live_results(cases, run)
    result = evaluate(cases, versions, run_mode=args.mode, run_metadata=run)
    write_reports(result, xlsx=args.xlsx)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    raise SystemExit(0 if result["status"] == "passed" else 1)


if __name__ == "__main__":
    main()
