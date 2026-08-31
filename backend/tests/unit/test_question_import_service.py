from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.models import (
    Base,
    DiagnosticQuestion,
    Domain,
    DomainChangeSet,
    KnowledgeDocument,
    KnowledgeItem,
)
from app.services import question_import_service
from app.services.question_import_service import (
    SHEET_QUESTIONS,
    build_question_template,
    create_import_run,
    question_gap_slots,
    publish_import_run,
    validate_import_run,
)
def _session():
    engine = create_engine("sqlite+pysqlite:///:memory:", poolclass=StaticPool)
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine, autoflush=False)()


def _workbook_with_questions(template: bytes) -> bytes:
    workbook = load_workbook(BytesIO(template))
    sheet = workbook[SHEET_QUESTIONS]
    columns = {cell.value: cell.column for cell in sheet[1]}
    for row in range(2, sheet.max_row + 1):
        purpose = sheet.cell(row, columns["purpose"]).value
        sheet.cell(row, columns["题目类型"]).value = "short_answer" if purpose == "mastery_validation" else "single_choice"
        sheet.cell(row, columns["难度"]).value = 2
        sheet.cell(row, columns["题干"]).value = "根据来源说明检索会召回与查询相关的知识切片。"
        sheet.cell(row, columns["解析"]).value = "来源说明检索会召回相关知识切片。"
        if purpose == "mastery_validation":
            sheet.cell(row, columns["正确答案"]).value = "检索会召回相关知识切片"
            sheet.cell(row, columns["评分点"]).value = "说明召回相关内容\n说明依据来源"
        else:
            for header, value in zip(["选项A", "选项B", "选项C", "选项D"], ["召回相关内容", "删除所有切片", "忽略查询", "替换来源"], strict=True):
                sheet.cell(row, columns[header]).value = value
            sheet.cell(row, columns["正确答案"]).value = "A"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_question_import_generates_only_missing_slots_and_publishes_atomically(monkeypatch) -> None:
    db = _session()
    db.add_all([
        Domain(domain_code="question_test", name="题库测试", status="preparing"),
        KnowledgeItem(
            public_id="ki_retrieval", domain_code="question_test", name="检索",
            category="RAG", difficulty=2, tags_json=["retrieval"],
            content_md="检索会根据查询召回与查询相关的知识切片，并保留来源。",
            source_title="测试资料", license_note="test", status="published",
        ),
    ])
    db.commit()
    template, _fingerprint, slot_count = build_question_template(db, "question_test")
    assert slot_count == 6
    assert len(question_gap_slots(db, "question_test")) == 6

    run = create_import_run(
        db,
        domain_code="question_test",
        original_name="questions.xlsx",
        content=_workbook_with_questions(template),
        created_by="tester",
    )
    assert run.status == "ready_to_publish"
    result = publish_import_run(db, run)
    assert result["status"] == "published"
    assert len(list(db.scalars(select(DiagnosticQuestion)))) == 6
    assert question_gap_slots(db, "question_test") == []


def test_question_import_can_persist_row_level_validation_progress(monkeypatch) -> None:
    db = _session()
    db.add_all([
        Domain(domain_code="question_progress", name="题库进度", status="preparing"),
        KnowledgeItem(
            public_id="ki_progress", domain_code="question_progress", name="检索",
            category="RAG", difficulty=2, tags_json=["retrieval"],
            content_md="检索会根据查询召回相关的知识切片，并保留来源。",
            source_title="测试资料", license_note="test", status="published",
        ),
    ])
    db.commit()
    template, _fingerprint, _slot_count = build_question_template(db, "question_progress")
    run = create_import_run(
        db,
        domain_code="question_progress",
        original_name="progress.xlsx",
        content=_workbook_with_questions(template),
        created_by="tester",
        validate_immediately=False,
    )
    assert run.status == "needs_attention"
    assert run.valid_row_count == 0

    result = validate_import_run(db, run, commit_progress=True)

    assert result.status == "ready_to_publish"
    assert result.valid_row_count == 6


def test_question_import_accepts_same_slot_name_for_different_knowledge_items(monkeypatch) -> None:
    db = _session()
    first = KnowledgeItem(
        public_id="ki_multi_first", domain_code="question_multi", name="检索基础",
        category="RAG", difficulty=2,
        content_md="检索会根据查询召回相关知识切片并保留来源。",
        source_title="测试资料", license_note="test", status="published",
    )
    second = KnowledgeItem(
        public_id="ki_multi_second", domain_code="question_multi", name="索引基础",
        category="RAG", difficulty=2,
        content_md="索引将知识切片组织为可检索的候选集合并保留来源。",
        source_title="测试资料", license_note="test", status="published",
    )
    db.add_all([Domain(domain_code="question_multi", name="题库测试"), first, second])
    db.commit()
    template, _fingerprint, slot_count = build_question_template(db, "question_multi")
    assert slot_count == 12
    create_import_run(
        db, domain_code="question_multi", original_name="multiple.xlsx",
        content=_workbook_with_questions(template), created_by="tester",
    )
    rows = list(db.scalars(select(question_import_service.QuestionImportRow)))
    assert all("SLOT_KEY_INVALID" not in row.validation_errors_json for row in rows)


def test_question_import_rejects_catalog_fingerprint_changes() -> None:
    db = _session()
    item = KnowledgeItem(
        public_id="ki_stale", domain_code="question_stale", name="旧知识", category="RAG",
        difficulty=2, content_md="旧知识内容", source_title="测试资料", license_note="test", status="published",
    )
    db.add_all([Domain(domain_code="question_stale", name="题库测试"), item])
    db.commit()
    template, _fingerprint, _slots = build_question_template(db, "question_stale")
    item.content_md = "已变化的知识内容"
    db.commit()

    run = create_import_run(
        db,
        domain_code="question_stale",
        original_name="questions.xlsx",
        content=_workbook_with_questions(template),
        created_by="tester",
    )
    assert run.status == "needs_attention"
    assert run.error_summary == "知识目录已变化，请重新下载模板"


def test_question_import_rejects_template_after_another_workbook_fills_its_slots(
    monkeypatch,
) -> None:
    db = _session()
    item = KnowledgeItem(
        public_id="ki_inventory", domain_code="question_inventory", name="检索",
        category="RAG", difficulty=2,
        content_md="检索会根据查询召回相关的知识切片，并保留来源。",
        source_title="测试资料", license_note="test", status="published",
    )
    db.add_all([Domain(domain_code="question_inventory", name="题库测试"), item])
    db.commit()
    template, _fingerprint, _slot_count = build_question_template(db, "question_inventory")
    first = create_import_run(
        db, domain_code="question_inventory", original_name="first.xlsx",
        content=_workbook_with_questions(template), created_by="tester",
    )
    publish_import_run(db, first)
    stale_book = load_workbook(BytesIO(_workbook_with_questions(template)))
    stale_sheet = stale_book[SHEET_QUESTIONS]
    stale_sheet.cell(2, {cell.value: cell.column for cell in stale_sheet[1]}["题干"]).value += "（旧版本）"
    stale_content = BytesIO()
    stale_book.save(stale_content)

    try:
        create_import_run(
            db, domain_code="question_inventory", original_name="stale.xlsx",
            content=stale_content.getvalue(), created_by="tester",
        )
    except question_import_service.QuestionImportError as exc:
        assert str(exc) == "QUESTION_TEMPLATE_INVENTORY_CHANGED"
    else:
        raise AssertionError("template must not overwrite slots published by another workbook")


def test_question_import_replaces_stale_questions_without_overwriting_history(monkeypatch) -> None:
    db = _session()
    item = KnowledgeItem(
        public_id="ki_replace", domain_code="question_replace", name="检索",
        category="RAG", difficulty=2,
        content_md="检索会根据查询召回相关知识切片并保留来源。",
        source_title="测试资料", license_note="test", status="published",
    )
    db.add_all([Domain(domain_code="question_replace", name="题库测试"), item])
    db.commit()
    first_template, _fingerprint, _slot_count = build_question_template(db, "question_replace")
    first = create_import_run(
        db, domain_code="question_replace", original_name="first.xlsx",
        content=_workbook_with_questions(first_template), created_by="tester",
    )
    publish_import_run(db, first)
    for question in db.scalars(select(DiagnosticQuestion)):
        question.status = "stale"
    db.commit()

    replacement_template, _fingerprint, slot_count = build_question_template(db, "question_replace")
    assert slot_count == 6
    replacement = create_import_run(
        db, domain_code="question_replace", original_name="replacement.xlsx",
        content=_workbook_with_questions(replacement_template), created_by="tester",
    )
    publish_import_run(db, replacement)

    questions = list(db.scalars(select(DiagnosticQuestion).order_by(DiagnosticQuestion.id)))
    assert len(questions) == 12
    assert sum(question.status == "disabled" for question in questions) == 6
    assert sum(question.status == "active" for question in questions) == 6
    assert all(
        ":replacement:" in str(question.external_id)
        for question in questions
        if question.status == "active"
    )


def test_change_set_question_import_stages_questions_until_activation(monkeypatch) -> None:
    db = _session()
    domain = Domain(domain_code="change_set_questions", name="变更集题库", status="ready")
    change_set = DomainChangeSet(
        public_id="dcs_questions", domain_code=domain.domain_code,
        status="ready_for_questions", mode="append", base_catalog_fingerprint="test",
    )
    db.add_all([domain, change_set])
    db.flush()
    document = KnowledgeDocument(
        public_id="kdoc_questions", domain_code=domain.domain_code,
        change_set_id=change_set.id, original_name="new.md", file_type="markdown",
        mime_type="text/markdown", sha256="a" * 64, source_title="测试资料",
        license_note="test", status="ready_for_questions",
    )
    db.add(document)
    db.flush()
    db.add(KnowledgeItem(
        public_id="ki_staged", domain_code=domain.domain_code, source_document_id=document.id,
        name="新增检索", category="RAG", difficulty=2,
        content_md="新增检索会根据查询召回相关的知识切片，并保留来源。",
        source_title="测试资料", license_note="test", status="staged",
    ))
    db.commit()
    template, _fingerprint, slot_count = build_question_template(
        db, domain.domain_code, change_set_id=change_set.public_id
    )
    assert slot_count == 6
    run = create_import_run(
        db, domain_code=domain.domain_code, original_name="staged.xlsx",
        content=_workbook_with_questions(template), created_by="tester",
        change_set_id=change_set.public_id,
    )
    result = publish_import_run(db, run)
    assert result["readiness"]["status"] == "staged_for_change_set"
    assert {row.status for row in db.scalars(select(DiagnosticQuestion))} == {"staged"}
    assert db.get(DomainChangeSet, change_set.id).status == "ready_to_activate"
