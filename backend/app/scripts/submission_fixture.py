"""Validate and load versioned competition submission fixtures."""

from __future__ import annotations

from collections import Counter
import hashlib
import json
from pathlib import Path
import re
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.agents.profile_analysis_config import MASTERY_BASELINES
from app.core.db import SessionLocal
from app.models import (
    DiagnosticQuestion,
    Domain,
    DomainIndexManifest,
    KnowledgeDocument,
    KnowledgeItem,
    KnowledgeRelation,
    Learner,
    LearnerProfile,
    LearningPath,
)
from app.rag.candidate_chunker import chunk_knowledge_item
from app.services.profile_service import build_learning_path_from_snapshot


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_FIXTURE_DIR = PROJECT_ROOT / "data" / "submission_fixtures" / "ai_app_dev_v1"
QUESTION_PURPOSES = {"diagnosis", "graded_quiz", "mastery_validation"}
PURPOSE_SLOTS = (
    ("diagnosis", "diagnosis_1", "foundation"),
    ("graded_quiz", "graded_foundation", "foundation"),
    ("graded_quiz", "graded_improvement", "improvement"),
    ("graded_quiz", "graded_challenge", "challenge"),
    ("mastery_validation", "mastery_1", "improvement"),
    ("mastery_validation", "mastery_2", "challenge"),
)


class SubmissionFixtureError(ValueError):
    pass


def _read_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise SubmissionFixtureError(f"invalid_fixture_json:{path.name}") from exc


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def resolve_fixture_dir(value: Path | str | None = None) -> Path:
    if value is None:
        return DEFAULT_FIXTURE_DIR
    value = Path(value)
    if value.is_absolute():
        return value
    candidates = (Path.cwd() / value, PROJECT_ROOT / value, Path("/app") / value)
    for candidate in candidates:
        if candidate.exists():
            return candidate.resolve()
    return candidates[0]


def _expected_relation_counts(counts: dict[str, Any]) -> Counter[str]:
    explicit = counts.get("relation_types")
    if isinstance(explicit, dict):
        return Counter({str(key): int(value) for key, value in explicit.items()})
    return Counter(
        {
            relation_type: int(counts[f"{relation_type}_relations"])
            for relation_type in ("prerequisite", "related")
            if f"{relation_type}_relations" in counts
        }
    )


def _validate_ai_app_dev_baseline(summary: dict[str, Any]) -> None:
    """Keep the first-domain submission baseline intentionally hash-strict."""
    counts = summary["counts"]
    if summary["fixture_version"] != "ai_app_dev_submission_fixture_v1":
        raise SubmissionFixtureError("fixture_ai_app_dev_version_invalid")
    if counts.get("knowledge_items") != 75 or counts.get("active_questions") != 465:
        raise SubmissionFixtureError("fixture_ai_app_dev_counts_invalid")
    if counts.get("question_purposes") != {
        "diagnosis": 90,
        "graded_quiz": 225,
        "mastery_validation": 150,
    }:
        raise SubmissionFixtureError("fixture_ai_app_dev_purpose_counts_invalid")
    if _expected_relation_counts(counts) != Counter({"prerequisite": 67, "related": 14}):
        raise SubmissionFixtureError("fixture_ai_app_dev_relation_counts_invalid")
    if counts.get("evaluation_cases") != 50 or counts.get("manual_demo_cases") != 3:
        raise SubmissionFixtureError("fixture_ai_app_dev_demo_counts_invalid")


def _validate_learner_profiles(
    payload: Any,
    *,
    domain_code: str,
    expected_count: int,
    knowledge_ids: set[str],
) -> list[dict[str, Any]]:
    if not isinstance(payload, list) or len(payload) != expected_count:
        raise SubmissionFixtureError("fixture_learner_profiles_count_invalid")
    learner_ids: set[str] = set()
    profile_ids: set[str] = set()
    path_ids: set[str] = set()
    for row in payload:
        if not isinstance(row, dict):
            raise SubmissionFixtureError("fixture_learner_profile_invalid")
        learner_id = row.get("learner_id")
        profile_id = row.get("profile_id")
        path_id = row.get("path_id")
        ability = row.get("ability_profile")
        weak = row.get("weak_knowledge")
        if (
            not all(isinstance(value, str) and value for value in (learner_id, profile_id, path_id))
            or learner_id in learner_ids
            or profile_id in profile_ids
            or path_id in path_ids
            or row.get("domain_code") != domain_code
            or not isinstance(ability, dict)
            or not isinstance(weak, list)
        ):
            raise SubmissionFixtureError("fixture_learner_profile_identity_invalid")
        required_scores = ("theory", "practice", "problem_solving", "breadth", "learning_speed")
        if not all(isinstance(ability.get(key), (int, float)) for key in required_scores):
            raise SubmissionFixtureError("fixture_learner_profile_ability_invalid")
        for weak_item in weak:
            if not isinstance(weak_item, dict) or weak_item.get("knowledge_id") not in knowledge_ids:
                raise SubmissionFixtureError("fixture_learner_profile_knowledge_invalid")
        learner_ids.add(learner_id)
        profile_ids.add(profile_id)
        path_ids.add(path_id)
    return payload


def _validate_smart_manufacturing_source_assets(
    root: Path,
    manifest: dict[str, Any],
    *,
    expected_knowledge: int,
) -> None:
    """Check the submitted XLSX metadata as well as its manifest hash."""
    assets = manifest.get("source_assets")
    if not isinstance(assets, dict):
        raise SubmissionFixtureError("fixture_source_assets_manifest_invalid")
    knowledge_path = root / str(assets.get("knowledge_package") or "")
    workbook_path = root / str(assets.get("question_workbook") or "")
    if not knowledge_path.is_file() or not workbook_path.is_file():
        raise SubmissionFixtureError("fixture_source_assets_missing")
    if len(re.findall(r"^## ", knowledge_path.read_text(encoding="utf-8"), flags=re.MULTILINE)) != expected_knowledge:
        raise SubmissionFixtureError("fixture_source_knowledge_count_invalid")
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "元数据" not in workbook.sheetnames:
            raise SubmissionFixtureError("fixture_source_workbook_metadata_missing")
        metadata = {
            str(key): str(value)
            for key, value in workbook["元数据"].iter_rows(min_row=1, max_col=2, values_only=True)
            if key is not None
        }
    finally:
        workbook.close()
    if (
        metadata.get("domain_code") != "smart_manufacturing"
        or metadata.get("knowledge_catalog_fingerprint") != assets.get("knowledge_catalog_fingerprint")
        or metadata.get("question_inventory_fingerprint") != assets.get("question_inventory_fingerprint")
    ):
        raise SubmissionFixtureError("fixture_source_workbook_fingerprint_invalid")


def validate_submission_fixture(fixture_dir: Path | str | None = None) -> dict[str, Any]:
    root = resolve_fixture_dir(fixture_dir)
    manifest_path = root / "manifest.json"
    manifest = _read_json(manifest_path)
    if not isinstance(manifest, dict):
        raise SubmissionFixtureError("fixture_manifest_not_object")
    if manifest.get("schema_version") != "submission-fixture-manifest-v1":
        raise SubmissionFixtureError("fixture_manifest_schema_invalid")
    version = manifest.get("fixture_version")
    domain_code = manifest.get("domain_code")
    files = manifest.get("files")
    counts = manifest.get("counts")
    if not isinstance(version, str) or not version or not isinstance(domain_code, str) or not domain_code:
        raise SubmissionFixtureError("fixture_manifest_identity_invalid")
    if not isinstance(files, dict) or not isinstance(counts, dict):
        raise SubmissionFixtureError("fixture_manifest_shape_invalid")
    for name, expected in files.items():
        if not isinstance(name, str) or not isinstance(expected, dict):
            raise SubmissionFixtureError("fixture_manifest_file_entry_invalid")
        path = root / name
        if not path.is_file() or _sha256(path) != expected.get("sha256"):
            raise SubmissionFixtureError(f"fixture_file_hash_mismatch:{name}")

    required_files = {
        "domain.json",
        "knowledge_items.json",
        "relations.json",
        "diagnostic_questions.json",
        "template_question_source.json",
        "import_source_manifest.json",
    }
    if not required_files.issubset(files):
        raise SubmissionFixtureError("fixture_required_files_missing")
    domain = _read_json(root / "domain.json")
    items = _read_json(root / "knowledge_items.json")
    relations = _read_json(root / "relations.json")
    questions = _read_json(root / "diagnostic_questions.json")
    template_source = _read_json(root / "template_question_source.json")
    supplemental = (
        _read_json(root / "supplemental_diagnosis_questions.json")
        if "supplemental_diagnosis_questions.json" in files
        else []
    )
    if not isinstance(domain, dict) or domain.get("domain_code") != domain_code:
        raise SubmissionFixtureError("fixture_domain_invalid")
    if not all(isinstance(value, list) for value in (items, relations, questions, template_source, supplemental)):
        raise SubmissionFixtureError("fixture_collection_invalid")

    expected_knowledge = counts.get("knowledge_items")
    knowledge_ids = [str(item.get("knowledge_id") or "") for item in items if isinstance(item, dict)]
    if (
        not isinstance(expected_knowledge, int)
        or len(items) != expected_knowledge
        or len(knowledge_ids) != expected_knowledge
        or len(set(knowledge_ids)) != expected_knowledge
        or not all(knowledge_ids)
    ):
        raise SubmissionFixtureError("fixture_knowledge_count_or_identity_invalid")
    if domain_code == "smart_manufacturing":
        _validate_smart_manufacturing_source_assets(
            root, manifest, expected_knowledge=expected_knowledge
        )
    knowledge_id_set = set(knowledge_ids)
    relation_counts: Counter[str] = Counter()
    relation_keys: set[tuple[str, str, str]] = set()
    for relation in relations:
        if not isinstance(relation, dict):
            raise SubmissionFixtureError("fixture_relation_invalid")
        relation_type = relation.get("relation_type")
        source = relation.get("source_knowledge_id")
        target = relation.get("target_knowledge_id")
        key = (str(relation_type), str(source), str(target))
        if (
            not isinstance(relation_type, str)
            or not relation_type
            or source not in knowledge_id_set
            or target not in knowledge_id_set
            or source == target
            or key in relation_keys
        ):
            raise SubmissionFixtureError("fixture_relation_reference_invalid")
        relation_keys.add(key)
        relation_counts[relation_type] += 1
    expected_relations = _expected_relation_counts(counts)
    if expected_relations and relation_counts != expected_relations:
        raise SubmissionFixtureError("fixture_relation_counts_invalid")
    if counts.get("knowledge_relations") not in (None, len(relations)):
        raise SubmissionFixtureError("fixture_relation_total_invalid")
    if domain_code == "smart_manufacturing" and any(
        relation.get("relation_type") != "next_step"
        or relation.get("generation_method") != "curriculum_rule"
        or dict(relation.get("evidence") or {}).get("evidence_kind") != "curriculum_rule"
        for relation in relations
    ):
        raise SubmissionFixtureError("fixture_curriculum_relation_evidence_invalid")

    question_ids: set[str] = set()
    external_ids: set[str] = set()
    purpose_counts: Counter[str] = Counter()
    coverage: dict[str, set[str]] = {purpose: set() for purpose in QUESTION_PURPOSES}
    for question in questions:
        if not isinstance(question, dict):
            raise SubmissionFixtureError("fixture_question_invalid")
        question_id = question.get("question_id")
        external_id = question.get("question_external_id")
        knowledge_id = question.get("knowledge_id")
        answer_key = question.get("answer_key")
        question_type = question.get("question_type")
        options = question.get("options")
        if not isinstance(question_id, str) or not question_id or question_id in question_ids:
            raise SubmissionFixtureError("fixture_question_id_invalid")
        if not isinstance(external_id, str) or not external_id or external_id in external_ids:
            raise SubmissionFixtureError("fixture_question_external_id_invalid")
        if knowledge_id not in knowledge_id_set or not isinstance(answer_key, dict):
            raise SubmissionFixtureError("fixture_question_reference_invalid")
        uses = answer_key.get("question_bank_uses")
        if not isinstance(uses, list) or len(uses) != 1 or uses[0] not in QUESTION_PURPOSES:
            raise SubmissionFixtureError("fixture_question_purpose_invalid")
        if not str(answer_key.get("explanation") or "").strip():
            raise SubmissionFixtureError("fixture_question_explanation_invalid")
        if question_type == "single_choice":
            correct_option = answer_key.get("correct_option")
            if not isinstance(options, list) or len(options) != 4 or not isinstance(correct_option, int) or not 0 <= correct_option < 4:
                raise SubmissionFixtureError("fixture_choice_question_invalid")
        elif question_type == "short_answer":
            if not isinstance(answer_key.get("rubric"), list) or not answer_key["rubric"]:
                raise SubmissionFixtureError("fixture_short_answer_invalid")
        else:
            raise SubmissionFixtureError("fixture_question_type_invalid")
        question_ids.add(question_id)
        external_ids.add(external_id)
        purpose = uses[0]
        purpose_counts[purpose] += 1
        coverage[purpose].add(knowledge_id)
    if len(questions) != counts.get("active_questions") or purpose_counts != Counter(counts.get("question_purposes") or {}):
        raise SubmissionFixtureError("fixture_question_counts_invalid")
    if any(values != knowledge_id_set for values in coverage.values()):
        raise SubmissionFixtureError("fixture_question_coverage_invalid")

    if len(template_source) != counts.get("template_compatible_questions") or len(supplemental) != int(counts.get("supplemental_diagnosis_questions") or 0):
        raise SubmissionFixtureError("fixture_template_source_counts_invalid")
    slot_keys = {(item.get("knowledge_id"), item.get("slot_key")) for item in template_source if isinstance(item, dict)}
    expected_slots = {(knowledge_id, slot_key) for knowledge_id in knowledge_id_set for _, slot_key, _ in PURPOSE_SLOTS}
    if len(slot_keys) != len(expected_slots) or slot_keys != expected_slots:
        raise SubmissionFixtureError("fixture_template_source_slot_coverage_invalid")
    for row in template_source:
        if not isinstance(row, dict) or row.get("source_question_id") not in question_ids:
            raise SubmissionFixtureError("fixture_template_source_reference_invalid")
        expected_purpose, _, expected_level = next(
            value for value in PURPOSE_SLOTS if value[1] == row.get("slot_key")
        )
        source_question = next(question for question in questions if question["question_id"] == row["source_question_id"])
        if (
            row.get("purpose") != expected_purpose
            or row.get("quiz_level") != expected_level
            or source_question["knowledge_id"] != row.get("knowledge_id")
            or source_question["answer_key"].get("question_bank_uses") != [expected_purpose]
        ):
            raise SubmissionFixtureError("fixture_template_source_purpose_invalid")
        if domain_code != "ai_app_dev" and source_question["answer_key"].get("quiz_level") != expected_level:
            raise SubmissionFixtureError("fixture_template_source_level_invalid")
    if len({row["source_question_id"] for row in template_source}) != len(template_source):
        raise SubmissionFixtureError("fixture_template_source_duplicate_reference")

    evaluation_count = int(counts.get("evaluation_cases") or 0)
    if evaluation_count:
        evaluation_cases = _read_json(root / "evaluation_cases_v4.json")
        if not isinstance(evaluation_cases, dict) or len(evaluation_cases.get("cases", [])) != evaluation_count:
            raise SubmissionFixtureError("fixture_evaluation_cases_invalid")
    elif "evaluation_cases_v4.json" in files:
        raise SubmissionFixtureError("fixture_unexpected_evaluation_cases")
    manual_count = int(counts.get("manual_demo_cases") or 0)
    if manual_count:
        manual_cases = _read_json(root / "manual_demo_cases.json")
        if not isinstance(manual_cases, dict) or len(manual_cases.get("cases", [])) != manual_count:
            raise SubmissionFixtureError("fixture_manual_cases_invalid")
    learner_count = int(counts.get("learner_profiles") or 0)
    learner_profiles: list[dict[str, Any]] = []
    if learner_count:
        learner_profiles = _validate_learner_profiles(
            _read_json(root / "learner_profiles.json"),
            domain_code=domain_code,
            expected_count=learner_count,
            knowledge_ids=knowledge_id_set,
        )

    import_source = manifest.get("import_source")
    if not isinstance(import_source, dict) or not isinstance(import_source.get("path"), str):
        raise SubmissionFixtureError("fixture_import_source_manifest_invalid")
    import_path = root / Path(import_source["path"])
    if not import_path.is_file() or _sha256(import_path) != import_source.get("sha256"):
        raise SubmissionFixtureError("fixture_import_source_hash_invalid")
    if len(re.findall(r"^## ", import_path.read_text(encoding="utf-8"), flags=re.MULTILINE)) != expected_knowledge:
        raise SubmissionFixtureError("fixture_import_source_knowledge_count_invalid")

    summary = {
        "status": "passed",
        "fixture_dir": str(root),
        "fixture_version": version,
        "fixture_sha256": _sha256(manifest_path),
        "domain_code": domain_code,
        "counts": {
            "knowledge_items": len(items),
            "knowledge_relations": len(relations),
            **({"prerequisite_relations": relation_counts["prerequisite"]} if "prerequisite_relations" in counts else {}),
            **({"related_relations": relation_counts["related"]} if "related_relations" in counts else {}),
            "active_questions": len(questions),
            "question_purposes": dict(sorted(purpose_counts.items())),
            "template_compatible_questions": len(template_source),
            "supplemental_diagnosis_questions": len(supplemental),
            **({"evaluation_cases": evaluation_count} if "evaluation_cases" in counts else {}),
            **({"manual_demo_cases": manual_count} if "manual_demo_cases" in counts else {}),
            **({"learner_profiles": len(learner_profiles)} if learner_count else {}),
        },
    }
    if domain_code == "ai_app_dev":
        _validate_ai_app_dev_baseline(summary)
    return summary


def _domain_config(payload: dict[str, Any], fixture_version: str, fixture_sha256: str) -> dict[str, Any]:
    profile_policy = dict(payload.get("profile_policy") or {})
    profile_policy.setdefault("version", f"{payload['domain_code']}_profile_v1")
    profile_policy.setdefault("ability_dimensions", ["theory", "practice", "problem_solving", "knowledge_breadth", "learning_speed"])
    profile_policy.setdefault("mastery_thresholds", [0.4, 0.6, 0.8])
    profile_policy.setdefault("mastery_baselines", MASTERY_BASELINES)
    profile_policy.setdefault("prior_mastery", 0.5)
    profile_policy.setdefault("prior_weight", 1.0)
    profile_policy.setdefault("minimum_effective_change", 5)
    profile_policy.setdefault("max_ability_change_per_update", 10)
    profile_policy.setdefault("max_weakness_level_change_per_update", 1)
    profile_policy.setdefault("default_n_results", 8)
    profile_policy.setdefault("multi_priority_remedial_n_results", 10)
    profile_policy.setdefault("maximum_n_results", 12)
    return {
        "resource_types": list(payload.get("resource_types") or []),
        "ability_dimensions": list(payload.get("ability_dimensions") or []),
        "learning_directions": list(payload.get("learning_directions") or []),
        "mvp_targets": dict(payload.get("mvp_targets") or {}),
        "readiness_policy": {"minimum_published_knowledge": 50, "minimum_diagnostic_questions": 60},
        "profile_policy": profile_policy,
        "submission_fixture": {"version": fixture_version, "sha256": fixture_sha256},
    }


def _fixture_state(db: Session, summary: dict[str, Any]) -> str:
    domains = list(db.scalars(select(Domain)))
    knowledge_count = len(list(db.scalars(select(KnowledgeItem))))
    question_count = len(list(db.scalars(select(DiagnosticQuestion))))
    if not domains and knowledge_count == 0 and question_count == 0:
        return "empty"
    if len(domains) != 1 or domains[0].domain_code != summary["domain_code"]:
        return "foreign"
    marker = dict(domains[0].config_json or {}).get("submission_fixture")
    if marker == {"version": summary["fixture_version"], "sha256": summary["fixture_sha256"]}:
        return "same_fixture"
    return "foreign"


def _assert_loaded_counts(db: Session, summary: dict[str, Any]) -> dict[str, Any]:
    domain_code = summary["domain_code"]
    items = list(db.scalars(select(KnowledgeItem).where(KnowledgeItem.domain_code == domain_code)))
    questions = list(db.scalars(select(DiagnosticQuestion).where(DiagnosticQuestion.domain_code == domain_code)))
    item_ids = {item.id for item in items}
    relations = [
        relation
        for relation in db.scalars(select(KnowledgeRelation))
        if relation.source_item_id in item_ids and relation.target_item_id in item_ids
    ]
    purposes = Counter(
        str((question.answer_key_json or {}).get("question_bank_uses", [""])[0])
        for question in questions
        if question.status == "active"
    )
    expected = summary["counts"]
    if len(items) != expected["knowledge_items"] or len(relations) != expected["knowledge_relations"] or len(questions) != expected["active_questions"] or purposes != Counter(expected["question_purposes"]):
        raise SubmissionFixtureError("loaded_fixture_counts_invalid")
    loaded = {
        "knowledge_items": len(items),
        "knowledge_relations": len(relations),
        "active_questions": len(questions),
        "question_purposes": dict(sorted(purposes.items())),
    }
    expected_profiles = int(expected.get("learner_profiles") or 0)
    if expected_profiles:
        profiles = list(db.scalars(select(LearnerProfile).where(LearnerProfile.domain_code == domain_code)))
        paths = list(db.scalars(select(LearningPath).where(LearningPath.domain_code == domain_code)))
        learners = list(db.scalars(select(Learner).where(Learner.target_domain == domain_code)))
        if len(profiles) != expected_profiles or len(paths) != expected_profiles or len(learners) != expected_profiles:
            raise SubmissionFixtureError("loaded_fixture_learner_profiles_invalid")
        loaded.update({"learners": len(learners), "learner_profiles": len(profiles), "learning_paths": len(paths)})
    index_manifest = db.scalar(select(DomainIndexManifest).where(DomainIndexManifest.domain_code == domain_code))
    loaded["index_version"] = index_manifest.index_version if index_manifest is not None else None
    return loaded


def _load_learner_profiles(
    db: Session,
    *,
    domain_code: str,
    profiles: list[dict[str, Any]],
) -> None:
    for fixture in profiles:
        learner = Learner(
            public_id=fixture["learner_id"],
            background=fixture["background"],
            education_level=fixture["education_level"],
            major=fixture["major"],
            target_domain=domain_code,
            experience_years=fixture["experience_years"],
            learning_style=fixture["learning_style"],
            direction_tags_json=fixture["direction_tags"],
            is_evaluation=True,
        )
        db.add(learner)
        db.flush()
        weak_knowledge = list(fixture["weak_knowledge"])
        profile = LearnerProfile(
            public_id=fixture["profile_id"],
            learner_id=learner.id,
            domain_code=domain_code,
            ability_profile_json=fixture["ability_profile"],
            weak_knowledge_json=weak_knowledge,
            profile_version=1,
            profile_source="submission_fixture",
            diagnosis_completed=True,
            changed_dimensions_json=["submission_fixture"],
            evidence_refs_json=[
                {
                    "evidence_id": f"submission_fixture:{fixture['learner_id']}",
                    "evidence_type": "synthetic_diagnostic_baseline",
                    "summary": "脱敏智能制造演示画像",
                    "confidence": 1.0,
                    "confirmed": True,
                }
            ],
            confidence=1.0,
            context_snapshot_json={
                "education_level": fixture["education_level"],
                "major": fixture["major"],
                "direction_tags": fixture["direction_tags"],
                "background": fixture["background"],
                "experience_years": fixture["experience_years"],
                "learning_style": fixture["learning_style"],
                "confirmed_at": fixture["confirmed_at"],
                "fixture_version": fixture["fixture_version"],
            },
            decision_reason=fixture["fixture_version"],
        )
        db.add(profile)
        db.flush()
        db.add(
            LearningPath(
                public_id=fixture["path_id"],
                learner_id=learner.id,
                profile_id=profile.id,
                domain_code=domain_code,
                status="active",
                path_json=build_learning_path_from_snapshot(fixture["ability_profile"], weak_knowledge),
                needs_refresh=False,
            )
        )


def load_submission_fixture(fixture_dir: Path | str | None = None) -> dict[str, Any]:
    summary = validate_submission_fixture(fixture_dir)
    root = Path(summary["fixture_dir"])
    with SessionLocal() as db:
        state = _fixture_state(db, summary)
        if state == "foreign":
            raise SubmissionFixtureError("fixture_requires_empty_database_or_same_fixture")
        if state == "same_fixture":
            return {**summary, "database": {"status": "already_loaded", **_assert_loaded_counts(db, summary)}}

        domain_payload = _read_json(root / "domain.json")
        item_payloads = _read_json(root / "knowledge_items.json")
        relation_payloads = _read_json(root / "relations.json")
        question_payloads = _read_json(root / "diagnostic_questions.json")
        profile_payloads = _read_json(root / "learner_profiles.json") if summary["counts"].get("learner_profiles") else []
        document = KnowledgeDocument(
            public_id=f"kdoc_{summary['fixture_version']}",
            domain_code=summary["domain_code"],
            original_name=f"{summary['fixture_version']}.json",
            stored_path=None,
            file_type="submission_fixture",
            mime_type="application/json",
            size_bytes=0,
            sha256=summary["fixture_sha256"],
            status="ready",
            knowledge_item_count=len(item_payloads),
            chunk_count=sum(
                len(chunk_knowledge_item(knowledge_id=item["knowledge_id"], name=item["name"], category=item["category"], difficulty=item["difficulty"], tags=item["tags"], content_md=item["content"]))
                for item in item_payloads
            ),
            source_title="比赛提交可执行测试数据夹具",
            license_note="Derived from the submitted knowledge package",
            uploaded_by="system",
        )
        db.add(document)
        domain = Domain(
            domain_code=summary["domain_code"],
            name=domain_payload["name"],
            status="ready",
            schema_version=domain_payload.get("domain_schema_version", "1.0"),
            config_json=_domain_config(domain_payload, summary["fixture_version"], summary["fixture_sha256"]),
        )
        db.add(domain)
        db.flush()
        item_by_public_id: dict[str, KnowledgeItem] = {}
        for item in item_payloads:
            model = KnowledgeItem(
                public_id=item["knowledge_id"],
                domain_code=item["domain_code"],
                name=item["name"],
                category=item["category"],
                difficulty=item["difficulty"],
                tags_json=item["tags"],
                evidence_capabilities_json=item["evidence_capabilities"],
                content_md=item["content"],
                source_title=item["source_title"],
                source_url=item["source_url"],
                license_note=item["license_note"],
                source_document_id=document.id,
                ability_weights_json=item["ability_weights"],
                status="published",
                needs_reembedding=True,
            )
            db.add(model)
            item_by_public_id[item["knowledge_id"]] = model
        db.flush()
        for relation in relation_payloads:
            db.add(
                KnowledgeRelation(
                    source_item_id=item_by_public_id[relation["source_knowledge_id"]].id,
                    target_item_id=item_by_public_id[relation["target_knowledge_id"]].id,
                    relation_type=relation["relation_type"],
                    evidence_json=dict(relation.get("evidence") or {"evidence_kind": "submission_fixture", "fixture_version": summary["fixture_version"]}),
                    generation_method=str(relation.get("generation_method") or "submission_fixture"),
                    source_document_id=document.id,
                )
            )
        for question in question_payloads:
            db.add(
                DiagnosticQuestion(
                    public_id=question["question_id"],
                    external_id=question["question_external_id"],
                    domain_code=summary["domain_code"],
                    knowledge_item_id=item_by_public_id[question["knowledge_id"]].id,
                    related_knowledge_ids_json=question["related_knowledge_ids"],
                    question_type=question["question_type"],
                    stem=question["stem"],
                    options_json=question["options"],
                    answer_key_json=question["answer_key"],
                    difficulty=question["difficulty"],
                    status="active",
                )
            )
        _load_learner_profiles(db, domain_code=summary["domain_code"], profiles=profile_payloads)
        db.commit()
        return {**summary, "database": {"status": "loaded", **_assert_loaded_counts(db, summary)}}
