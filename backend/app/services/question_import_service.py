"""Independent, spreadsheet-driven formal question-bank import.

Document imports publish knowledge assets only.  This module turns a generated
gap template into staged formal questions without coupling question authoring to
the document-import worker or its model-generation batches.
"""

from __future__ import annotations

import hashlib
import json
import logging
import re
from threading import Lock, Thread
from datetime import UTC, datetime
from io import BytesIO
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.core.db import SessionLocal
from app.models import (
    DiagnosticQuestion,
    Domain,
    DomainChangeSet,
    KnowledgeDocument,
    KnowledgeItem,
    QuestionImportRow,
    QuestionImportRun,
)
from app.rag.candidate_chunker import CHUNKER_VERSION
from app.services.domain_api_service import DomainApiService
from app.services.question_bank_service import question_bank_coverage
from app.services.question_certification_service import (
    QUESTION_CERTIFICATION_RULE_VERSION,
    certify_question_payloads,
    knowledge_item_content_hash,
    normalize_evidence_text,
)
from app.services.question_source_binding_service import candidate_chunks_for_item, candidate_source_locator


TEMPLATE_VERSION = "question-bank-xlsx-v2"
SHEET_QUESTIONS = "题目"
SHEET_GUIDE = "填写说明"
SHEET_ENUMS = "枚举"
SHEET_META = "元数据"
HEADERS = [
    "slot_key",
    "question_external_id",
    "knowledge_ref",
    "知识点名称",
    "purpose",
    "quiz_level",
    "domain_code",
    "knowledge_catalog_fingerprint",
    "题目类型",
    "难度",
    "题干",
    "选项A",
    "选项B",
    "选项C",
    "选项D",
    "正确答案",
    "解析",
    "评分点",
]
LOCKED_COLUMNS = {"A", "B", "C", "D", "E", "F", "G", "H"}
EDITABLE_COLUMNS = {"I", "J", "K", "L", "M", "N", "O", "P", "Q", "R"}
PURPOSE_SLOTS = {
    "diagnosis": [("diagnosis_1", "foundation")],
    "graded_quiz": [
        ("graded_foundation", "foundation"),
        ("graded_improvement", "improvement"),
        ("graded_challenge", "challenge"),
    ],
    "mastery_validation": [
        ("mastery_1", "improvement"),
        ("mastery_2", "challenge"),
    ],
}
VALID_PURPOSES = set(PURPOSE_SLOTS)
VALID_TYPES = {"single_choice", "short_answer"}
VALID_LEVELS = {"foundation", "improvement", "challenge"}
logger = logging.getLogger(__name__)
_validation_lock = Lock()
_active_validations: set[str] = set()


class QuestionImportError(ValueError):
    pass


def _attention_summary(rows: list[QuestionImportRow]) -> str | None:
    labels = (
        ("source_confirmation_required", "待确认来源"),
        ("content_rejected", "内容不合格"),
        ("certification_service_error", "认证异常"),
        ("template_invalid", "模板字段错误"),
    )
    parts = [
        f"{sum(row.status == status for row in rows)} 题{label}"
        for status, label in labels
        if any(row.status == status for row in rows)
    ]
    return "；".join(parts) or None


def _text(value: object) -> str:
    return str(value or "").strip()


def _catalog_items(
    db: Session, domain_code: str, *, change_set_id: int | None = None
) -> list[KnowledgeItem]:
    """Active catalog plus newly staged knowledge in one pending change set."""
    filters = [
        KnowledgeItem.domain_code == domain_code,
        KnowledgeItem.status == "published",
    ]
    if change_set_id is not None:
        staged_document_ids = select(KnowledgeDocument.id).where(
            KnowledgeDocument.change_set_id == change_set_id,
            KnowledgeDocument.domain_code == domain_code,
        )
        filters = [
            KnowledgeItem.domain_code == domain_code,
            or_(
                KnowledgeItem.status == "published",
                (KnowledgeItem.status == "staged")
                & KnowledgeItem.source_document_id.in_(staged_document_ids),
            ),
        ]
    return list(db.scalars(select(KnowledgeItem).where(*filters).order_by(KnowledgeItem.public_id)))


def knowledge_catalog_fingerprint(items: list[KnowledgeItem]) -> str:
    payload = [
        {
            "knowledge_ref": item.public_id,
            "content_hash": knowledge_item_content_hash(item),
            "difficulty": item.difficulty,
        }
        for item in items
    ]
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return "sha256:" + hashlib.sha256(encoded.encode()).hexdigest()


def _slot_external_id(domain_code: str, knowledge_ref: str, slot_key: str) -> str:
    return f"qslot:{domain_code}:{knowledge_ref}:{slot_key}"


def _replacement_external_id(
    db: Session, *, domain_code: str, knowledge_ref: str, slot_key: str
) -> str:
    """Keep prior attempts immutable when a disabled imported slot is replaced."""

    base = _slot_external_id(domain_code, knowledge_ref, slot_key)
    existing = set(
        db.scalars(
            select(DiagnosticQuestion.external_id).where(
                DiagnosticQuestion.domain_code == domain_code,
                DiagnosticQuestion.external_id.like(f"{base}%"),
            )
        )
    )
    if base not in existing:
        return base
    revision = 2
    while f"{base}:replacement:{revision}" in existing:
        revision += 1
    return f"{base}:replacement:{revision}"


def question_gap_slots(
    db: Session,
    domain_code: str,
    *,
    knowledge_refs: set[str] | None = None,
    change_set_id: int | None = None,
) -> list[dict[str, str]]:
    items = _catalog_items(db, domain_code, change_set_id=change_set_id)
    if knowledge_refs is not None:
        items = [item for item in items if item.public_id in knowledge_refs]
    coverage = question_bank_coverage(
        db, domain_code=domain_code, knowledge_ids=[item.public_id for item in items]
    )
    rows = list(
        db.execute(
            select(DiagnosticQuestion, KnowledgeItem.public_id)
            .join(KnowledgeItem, DiagnosticQuestion.knowledge_item_id == KnowledgeItem.id)
            .where(
                DiagnosticQuestion.domain_code == domain_code,
                DiagnosticQuestion.status == "active",
                DiagnosticQuestion.certification_status == "certified",
                KnowledgeItem.status == "published",
            )
        )
    )
    existing_levels: dict[tuple[str, str], list[str]] = {}
    for question, knowledge_ref in rows:
        purpose = next(iter((question.answer_key_json or {}).get("question_bank_uses") or []), "")
        if purpose in VALID_PURPOSES:
            existing_levels.setdefault((knowledge_ref, purpose), []).append(
                _text((question.answer_key_json or {}).get("quiz_level"))
            )
    if change_set_id is not None:
        # Staged workbooks are part of the pending target inventory.  They are
        # deliberately invisible to learner selection until the change set is
        # activated, but later batches must not receive duplicate slots.
        staged_questions = list(
            db.scalars(
                select(DiagnosticQuestion).where(
                    DiagnosticQuestion.domain_code == domain_code,
                    DiagnosticQuestion.status == "staged",
                    DiagnosticQuestion.certification_status == "certified",
                )
            )
        )
        item_by_id = {item.id: item.public_id for item in items}
        for question in staged_questions:
            answer_key = dict(question.answer_key_json or {})
            if str(answer_key.get("pending_change_set_id") or "") != str(change_set_id):
                continue
            knowledge_ref = item_by_id.get(question.knowledge_item_id)
            purpose = next(iter(answer_key.get("question_bank_uses") or []), "")
            if knowledge_ref is None or purpose not in VALID_PURPOSES:
                continue
            values = coverage["counts_by_knowledge"].setdefault(knowledge_ref, {
                "single_choice": 0, "short_answer": 0, "total": 0,
                "diagnosis": 0, "graded_quiz": 0, "mastery_validation": 0,
                "mastery_reserve": 0,
            })
            values[purpose] += 1
            values["total"] += 1
            values[question.question_type] += 1
            existing_levels.setdefault((knowledge_ref, purpose), []).append(
                _text(answer_key.get("quiz_level"))
            )
    result: list[dict[str, str]] = []
    counts = coverage["counts_by_knowledge"]
    for item in items:
        values = counts[item.public_id]
        for purpose, slots in PURPOSE_SLOTS.items():
            required = len(slots)
            available = int(values[purpose])
            if available >= required:
                continue
            used_levels = list(existing_levels.get((item.public_id, purpose), []))
            missing = required - available
            for slot_key, level in slots:
                if missing <= 0:
                    break
                if purpose == "graded_quiz" and level in used_levels:
                    continue
                result.append(
                    {
                        "slot_key": slot_key,
                        "question_external_id": _replacement_external_id(
                            db,
                            domain_code=domain_code,
                            knowledge_ref=item.public_id,
                            slot_key=slot_key,
                        ),
                        "knowledge_ref": item.public_id,
                        "knowledge_name": item.name,
                        "purpose": purpose,
                        "quiz_level": level,
                    }
                )
                missing -= 1
    return result


def question_inventory_fingerprint(slots: list[dict[str, str]]) -> str:
    """Fingerprint exact missing slots so concurrent workbooks cannot both publish."""
    payload = [
        {
            "slot_key": slot["slot_key"],
            "question_external_id": slot["question_external_id"],
            "knowledge_ref": slot["knowledge_ref"],
            "purpose": slot["purpose"],
            "quiz_level": slot["quiz_level"],
        }
        for slot in sorted(
            slots,
            key=lambda value: (
                value["knowledge_ref"], value["purpose"], value["slot_key"],
                value["question_external_id"],
            ),
        )
    ]
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return "sha256:" + hashlib.sha256(encoded.encode()).hexdigest()


def _scope_text(knowledge_refs: set[str] | None) -> str:
    return ",".join(sorted(knowledge_refs or set()))


def _scope_from_text(value: str) -> set[str] | None:
    refs = {item.strip() for item in value.split(",") if item.strip()}
    return refs or None


def build_question_template(
    db: Session,
    domain_code: str,
    *,
    knowledge_refs: set[str] | None = None,
    change_set_id: str | None = None,
) -> tuple[bytes, str, int]:
    change_set = None
    if change_set_id:
        change_set = db.scalar(
            select(DomainChangeSet).where(
                DomainChangeSet.public_id == change_set_id,
                DomainChangeSet.domain_code == domain_code,
            )
        )
        if change_set is None:
            raise QuestionImportError("DOMAIN_CHANGE_SET_NOT_FOUND")
        if change_set.status not in {"preparing", "ready_for_questions", "questions_preparing"}:
            raise QuestionImportError("DOMAIN_CHANGE_SET_NOT_ACCEPTING_QUESTIONS")
    items = _catalog_items(
        db, domain_code, change_set_id=change_set.id if change_set else None
    )
    if knowledge_refs is not None:
        unknown_refs = knowledge_refs - {item.public_id for item in items}
        if unknown_refs:
            raise QuestionImportError("QUESTION_TEMPLATE_KNOWLEDGE_SCOPE_INVALID")
        items = [item for item in items if item.public_id in knowledge_refs]
    if not items:
        raise QuestionImportError("DOMAIN_HAS_NO_PUBLISHED_KNOWLEDGE")
    fingerprint = knowledge_catalog_fingerprint(items)
    slots = question_gap_slots(
        db,
        domain_code,
        knowledge_refs=knowledge_refs,
        change_set_id=change_set.id if change_set else None,
    )
    inventory = question_inventory_fingerprint(slots)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_QUESTIONS
    sheet.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.protection = Protection(locked=True)
    for slot in slots:
        sheet.append(
            [
                slot["slot_key"], slot["question_external_id"], slot["knowledge_ref"],
                slot["knowledge_name"], slot["purpose"], slot["quiz_level"], domain_code,
                fingerprint, "", "", "", "", "", "", "", "", "", "",
            ]
        )
    for row in sheet.iter_rows(min_row=2, max_row=max(2, sheet.max_row)):
        for cell in row:
            cell.protection = Protection(locked=cell.column_letter not in EDITABLE_COLUMNS)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:R{max(1, sheet.max_row)}"
    for column, width in {
        "A": 20, "B": 38, "C": 24, "D": 24, "E": 20, "F": 18, "G": 18,
        "H": 72, "I": 18, "J": 10, "K": 44, "L": 22, "M": 22, "N": 22,
        "O": 22, "P": 18, "Q": 44, "R": 36,
    }.items():
        sheet.column_dimensions[column].width = width
    for column in LOCKED_COLUMNS:
        sheet.column_dimensions[column].hidden = column in {"A", "B", "C", "G", "H"}
    sheet.protection.sheet = True
    sheet.protection.set_password("cognivia-template")

    enums = workbook.create_sheet(SHEET_ENUMS)
    enums.append(["题目类型", "难度", "答案示例"])
    enums.append(["single_choice", 1, "A"])
    enums.append(["short_answer", 2, "填写参考答案"])
    for value in range(3, 6):
        enums.append([None, value, None])
    enums.sheet_state = "hidden"
    type_validation = DataValidation(type="list", formula1=f"'{SHEET_ENUMS}'!$A$2:$A$3")
    difficulty_validation = DataValidation(type="list", formula1=f"'{SHEET_ENUMS}'!$B$2:$B$6")
    sheet.add_data_validation(type_validation)
    sheet.add_data_validation(difficulty_validation)
    type_validation.add(f"I2:I{max(2, sheet.max_row)}")
    difficulty_validation.add(f"J2:J{max(2, sheet.max_row)}")
    sheet.conditional_formatting.add(
        f"I2:I{max(2, sheet.max_row)}",
        FormulaRule(formula=["LEN(I2)=0"], fill=PatternFill("solid", fgColor="FFF2CC")),
    )

    guide = workbook.create_sheet(SHEET_GUIDE)
    guide.append(["填写规则"])
    for text in [
        "系统已安排知识点、用途和题目槽位；请勿修改受保护字段。",
        "single_choice 必须填写 4 个不重复选项，正确答案填 A、B、C 或 D。",
        "short_answer 不填写选项，正确答案填写参考答案，评分点用换行分隔并填写 2-8 条。",
        "系统会自动绑定来源；仅未能确认来源的行需要在管理页选择并圈选原文。",
    ]:
        guide.append([text])
    guide.column_dimensions["A"].width = 110

    meta = workbook.create_sheet(SHEET_META)
    for key, value in [
        ("template_version", TEMPLATE_VERSION),
        ("domain_code", domain_code),
        ("knowledge_catalog_fingerprint", fingerprint),
        ("question_inventory_fingerprint", inventory),
        ("knowledge_scope", _scope_text(knowledge_refs)),
        ("change_set_id", change_set.public_id if change_set else ""),
        ("generated_at", datetime.now(UTC).isoformat()),
    ]:
        meta.append([key, value])
    meta.sheet_state = "hidden"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), fingerprint, len(slots)


def _metadata(workbook) -> dict[str, str]:
    if SHEET_META not in workbook.sheetnames:
        raise QuestionImportError("QUESTION_TEMPLATE_METADATA_MISSING")
    return {
        _text(row[0]): _text(row[1])
        for row in workbook[SHEET_META].iter_rows(min_row=1, values_only=True)
        if len(row) >= 2 and _text(row[0])
    }


def _rows_from_workbook(content: bytes) -> tuple[dict[str, str], list[tuple[int, dict[str, Any]]]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise QuestionImportError("QUESTION_TEMPLATE_XLSX_INVALID") from exc
    if SHEET_QUESTIONS not in workbook.sheetnames:
        raise QuestionImportError("QUESTION_TEMPLATE_SHEET_MISSING")
    sheet = workbook[SHEET_QUESTIONS]
    headers = [_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    if headers != HEADERS:
        raise QuestionImportError("QUESTION_TEMPLATE_HEADERS_INVALID")
    rows: list[tuple[int, dict[str, Any]]] = []
    for number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {header: values[index] if index < len(values) else None for index, header in enumerate(HEADERS)}
        if not any(_text(value) for value in record.values()):
            continue
        rows.append((number, record))
    if not rows:
        raise QuestionImportError("QUESTION_TEMPLATE_HAS_NO_ROWS")
    return _metadata(workbook), rows


def _normalize_row(record: dict[str, Any]) -> dict[str, Any]:
    question_type = _text(record["题目类型"])
    options = [_text(record[column]) for column in ("选项A", "选项B", "选项C", "选项D")]
    purpose = _text(record["purpose"])
    answer = _text(record["正确答案"])
    if question_type == "single_choice":
        answer_value: object = ord(answer.upper()) - ord("A") if answer.upper() in {"A", "B", "C", "D"} else answer
    else:
        answer_value = answer
    return {
        "template_domain_code": _text(record["domain_code"]),
        "template_knowledge_catalog_fingerprint": _text(
            record["knowledge_catalog_fingerprint"]
        ),
        "knowledge_candidate_id": _text(record["knowledge_ref"]),
        "question_slot": _text(record["slot_key"]),
        "question_bank_uses": [purpose],
        "quiz_level": _text(record["quiz_level"]),
        "question_type": question_type,
        "difficulty": record["难度"],
        "stem": _text(record["题干"]),
        "options": options if question_type == "single_choice" else [],
        "answer": answer_value,
        "explanation": _text(record["解析"]),
        "rubric": [item.strip() for item in re.split(r"[\n|]", _text(record["评分点"])) if item.strip()],
    }


def _candidate_sources(item: KnowledgeItem, payload: dict[str, Any]) -> list[dict[str, str]]:
    needle = normalize_evidence_text(
        " ".join(
            [payload.get("stem", ""), str(payload.get("answer", "")), payload.get("explanation", "")]
        )
    )
    scored: list[tuple[float, Any]] = []
    for chunk in candidate_chunks_for_item(item):
        source = normalize_evidence_text(chunk.content)
        overlap = len(set(needle) & set(source)) / max(1, len(set(needle)))
        scored.append((overlap, chunk))
    scored.sort(key=lambda value: (-value[0], value[1].chunk_index))
    return [
        {
            "source_ref_id": chunk.chunk_id,
            "source_locator": candidate_source_locator(item, chunk),
            "knowledge_id": item.public_id,
            "excerpt": chunk.content[:500],
            "score": round(score, 4),
        }
        for score, chunk in scored[:3]
    ]


def _automatic_binding(item: KnowledgeItem, candidates: list[dict[str, str]]) -> dict[str, Any] | None:
    if not candidates or float(candidates[0]["score"]) < 0.45:
        return None
    candidate = candidates[0]
    chunks = {chunk.chunk_id: chunk for chunk in candidate_chunks_for_item(item)}
    chunk = chunks.get(candidate["source_ref_id"])
    if chunk is None:
        return None
    quote = chunk.content[: min(320, len(chunk.content))].strip()
    return {"source_ref_ids": [chunk.chunk_id], "quotes": {chunk.chunk_id: quote}}


def _binding_payload(item: KnowledgeItem, payload: dict[str, Any], binding: dict[str, Any]) -> dict[str, Any]:
    chunks = {chunk.chunk_id: chunk for chunk in candidate_chunks_for_item(item)}
    source_ref_ids = [str(value) for value in binding.get("source_ref_ids") or []]
    quotes = dict(binding.get("quotes") or {})
    if not 1 <= len(source_ref_ids) <= 3 or len(source_ref_ids) != len(set(source_ref_ids)):
        raise QuestionImportError("QUESTION_SOURCE_SELECTION_INVALID")
    max_sources = {"foundation": 1, "improvement": 2, "challenge": 3}.get(
        str(payload.get("quiz_level")), 0
    )
    if len(source_ref_ids) > max_sources:
        raise QuestionImportError("QUESTION_SOURCE_SELECTION_EXCEEDS_LEVEL")
    source_chunks = []
    evidence_quotes = []
    source_hashes = {}
    for source_ref_id in source_ref_ids:
        chunk = chunks.get(source_ref_id)
        quote = _text(quotes.get(source_ref_id))
        if chunk is None or not quote or normalize_evidence_text(quote) not in normalize_evidence_text(chunk.content):
            raise QuestionImportError("QUESTION_SOURCE_QUOTE_INVALID")
        locator = candidate_source_locator(item, chunk)
        source_chunks.append(
            {
                "chunk_id": source_ref_id,
                "chunk_index": chunk.chunk_index,
                "source_locator": locator,
                "knowledge_candidate_id": item.public_id,
                "knowledge_id": item.public_id,
                "content": chunk.content,
                "source_content_hash": knowledge_item_content_hash(item),
                "chunker_version": CHUNKER_VERSION,
            }
        )
        evidence_quotes.append({"source_ref_id": source_ref_id, "quote": quote})
        source_hashes[source_ref_id] = knowledge_item_content_hash(item)
    aggregate_hash = "sha256:" + hashlib.sha256(
        json.dumps(source_hashes, sort_keys=True, separators=(",", ":")).encode()
    ).hexdigest()
    return {
        **payload,
        "source_quote": evidence_quotes[0]["quote"],
        "source_chunks": source_chunks,
        "source_ref_ids": source_ref_ids,
        "source_locators": {item_["chunk_id"]: item_["source_locator"] for item_ in source_chunks},
        "source_content_hashes": source_hashes,
        "source_content_hash": aggregate_hash,
        "evidence_quotes": evidence_quotes,
        "chunker_version": CHUNKER_VERSION,
    }


def create_import_run(
    db: Session,
    *,
    domain_code: str,
    original_name: str,
    content: bytes,
    created_by: str,
    change_set_id: str | None = None,
    validate_immediately: bool = True,
) -> QuestionImportRun:
    if not original_name.lower().endswith(".xlsx"):
        raise QuestionImportError("QUESTION_IMPORT_XLSX_REQUIRED")
    file_sha256 = hashlib.sha256(content).hexdigest()
    existing = db.scalar(
        select(QuestionImportRun).where(
            QuestionImportRun.domain_code == domain_code,
            QuestionImportRun.file_sha256 == file_sha256,
            QuestionImportRun.status != "cancelled",
        )
    )
    if existing is not None:
        return existing
    metadata, raw_rows = _rows_from_workbook(content)
    if metadata.get("template_version") != TEMPLATE_VERSION or metadata.get("domain_code") != domain_code:
        raise QuestionImportError("QUESTION_TEMPLATE_DOMAIN_OR_VERSION_INVALID")
    change_set = None
    template_change_set_id = metadata.get("change_set_id", "")
    if (change_set_id or "") != template_change_set_id:
        raise QuestionImportError("QUESTION_TEMPLATE_CHANGE_SET_INVALID")
    if change_set_id:
        change_set = db.scalar(
            select(DomainChangeSet).where(
                DomainChangeSet.public_id == change_set_id,
                DomainChangeSet.domain_code == domain_code,
            )
        )
        if change_set is None:
            raise QuestionImportError("DOMAIN_CHANGE_SET_NOT_FOUND")
        if change_set.status not in {"preparing", "ready_for_questions", "questions_preparing"}:
            raise QuestionImportError("DOMAIN_CHANGE_SET_NOT_ACCEPTING_QUESTIONS")
    items = _catalog_items(
        db, domain_code, change_set_id=change_set.id if change_set else None
    )
    fingerprint = knowledge_catalog_fingerprint(items)
    knowledge_scope = _scope_from_text(metadata.get("knowledge_scope", ""))
    if knowledge_scope is not None and not knowledge_scope <= {item.public_id for item in items}:
        raise QuestionImportError("QUESTION_TEMPLATE_KNOWLEDGE_SCOPE_INVALID")
    expected_slots = question_gap_slots(
        db,
        domain_code,
        knowledge_refs=knowledge_scope,
        change_set_id=change_set.id if change_set else None,
    )
    inventory = question_inventory_fingerprint(expected_slots)
    if metadata.get("question_inventory_fingerprint") != inventory:
        raise QuestionImportError("QUESTION_TEMPLATE_INVENTORY_CHANGED")
    run = QuestionImportRun(
        public_id=f"qir_{uuid4().hex[:16]}",
        domain_code=domain_code,
        template_version=TEMPLATE_VERSION,
        knowledge_catalog_fingerprint=metadata.get("knowledge_catalog_fingerprint", ""),
        question_inventory_fingerprint=metadata.get("question_inventory_fingerprint", ""),
        scope_json=sorted(knowledge_scope or []),
        change_set_id=change_set.id if change_set else None,
        original_name=original_name[:255],
        file_sha256=file_sha256,
        status="uploaded",
        row_count=len(raw_rows),
        valid_row_count=0,
        created_by=created_by[:64],
    )
    db.add(run)
    db.flush()
    for number, record in raw_rows:
        db.add(
            QuestionImportRow(
                public_id=f"qrow_{uuid4().hex[:16]}",
                run_id=run.id,
                row_number=number,
                question_external_id=_text(record["question_external_id"]),
                slot_key=_text(record["slot_key"]),
                knowledge_ref=_text(record["knowledge_ref"]),
                payload_json=_normalize_row(record),
                candidate_sources_json=[],
                source_binding_json={},
                certification_report_json={},
                status="pending",
                validation_errors_json=[],
            )
        )
    db.flush()
    if validate_immediately:
        validate_import_run(db, run, current_fingerprint=fingerprint)
    else:
        run.status = "validating"
        run.error_summary = None
    db.commit()
    db.refresh(run)
    return run


def validate_import_run(
    db: Session,
    run: QuestionImportRun,
    *,
    current_fingerprint: str | None = None,
    commit_progress: bool = False,
) -> QuestionImportRun:
    items = _catalog_items(db, run.domain_code, change_set_id=run.change_set_id)
    current = current_fingerprint or knowledge_catalog_fingerprint(items)
    rows = list(db.scalars(select(QuestionImportRow).where(QuestionImportRow.run_id == run.id)))
    if run.knowledge_catalog_fingerprint != current:
        for row in rows:
            row.status = "template_invalid"
            row.validation_errors_json = ["KNOWLEDGE_CATALOG_CHANGED"]
        run.status = "needs_attention"
        run.error_summary = "知识目录已变化，请重新下载模板"
        run.valid_row_count = 0
        if commit_progress:
            db.commit()
        else:
            db.flush()
        return run
    knowledge_scope = {str(value) for value in (run.scope_json or []) if str(value)} or None
    expected_slots = question_gap_slots(
        db,
        run.domain_code,
        knowledge_refs=knowledge_scope,
        change_set_id=run.change_set_id,
    )
    expected_inventory = question_inventory_fingerprint(expected_slots)
    if run.question_inventory_fingerprint != expected_inventory:
        for row in rows:
            row.status = "template_invalid"
            row.validation_errors_json = ["QUESTION_INVENTORY_CHANGED"]
        run.status = "needs_attention"
        run.error_summary = "题库缺口已变化，请重新下载模板"
        run.valid_row_count = 0
        if commit_progress:
            db.commit()
        else:
            db.flush()
        return run
    expected_row_keys = {
        (slot["knowledge_ref"], slot["slot_key"], slot["question_external_id"])
        for slot in expected_slots
    }
    row_keys = {(row.knowledge_ref, row.slot_key, row.question_external_id) for row in rows}
    if row_keys != expected_row_keys:
        for row in rows:
            row.status = "template_invalid"
            row.validation_errors_json = ["QUESTION_TEMPLATE_SLOT_SET_INVALID"]
        run.status = "needs_attention"
        run.error_summary = "模板题目槽位不完整或已被修改，请重新下载模板"
        run.valid_row_count = 0
        if commit_progress:
            db.commit()
        else:
            db.flush()
        return run
    item_by_ref = {item.public_id: item for item in items}
    seen_slots: set[tuple[str, str]] = set()
    valid_count = 0
    eligible: list[tuple[str, dict[str, Any]]] = []
    row_by_id = {row.public_id: row for row in rows}
    run.status = "validating"
    run.error_summary = None
    for row in rows:
        payload = dict(row.payload_json or {})
        errors: list[str] = []
        item = item_by_ref.get(row.knowledge_ref)
        slot_identity = (row.knowledge_ref, row.slot_key)
        if not row.question_external_id or not row.slot_key or slot_identity in seen_slots:
            errors.append("SLOT_KEY_INVALID")
        seen_slots.add(slot_identity)
        if item is None:
            errors.append("KNOWLEDGE_REF_INVALID")
        if (
            _text(payload.get("template_domain_code")) != run.domain_code
            or _text(payload.get("template_knowledge_catalog_fingerprint"))
            != run.knowledge_catalog_fingerprint
        ):
            errors.append("QUESTION_TEMPLATE_METADATA_INVALID")
        purpose = next(iter(payload.get("question_bank_uses") or []), "")
        if purpose not in VALID_PURPOSES:
            errors.append("PURPOSE_INVALID")
        else:
            expected_slots = dict(PURPOSE_SLOTS[purpose])
            base_external_id = _slot_external_id(
                run.domain_code, row.knowledge_ref, row.slot_key
            )
            if (
                row.slot_key not in expected_slots
                or _text(payload.get("quiz_level")) != expected_slots[row.slot_key]
                or not (
                    row.question_external_id == base_external_id
                    or row.question_external_id.startswith(f"{base_external_id}:replacement:")
                )
            ):
                errors.append("TEMPLATE_SLOT_METADATA_INVALID")
        if payload.get("question_type") not in VALID_TYPES:
            errors.append("QUESTION_TYPE_INVALID")
        try:
            if not 1 <= int(payload.get("difficulty")) <= 5:
                errors.append("DIFFICULTY_INVALID")
            else:
                payload["difficulty"] = int(payload["difficulty"])
        except (TypeError, ValueError):
            errors.append("DIFFICULTY_INVALID")
        if not _text(payload.get("stem")) or not _text(payload.get("explanation")):
            errors.append("QUESTION_CONTENT_REQUIRED")
        if payload.get("question_type") == "single_choice":
            options = list(payload.get("options") or [])
            if len(options) != 4 or not all(options) or len(set(options)) != 4 or not isinstance(payload.get("answer"), int) or not 0 <= payload["answer"] < 4:
                errors.append("SINGLE_CHOICE_VALUES_INVALID")
        if payload.get("question_type") == "short_answer" and (
            not _text(payload.get("answer")) or not 2 <= len(payload.get("rubric") or []) <= 8
        ):
            errors.append("SHORT_ANSWER_VALUES_INVALID")
        if item is not None:
            candidates = _candidate_sources(item, payload)
            row.candidate_sources_json = candidates
            binding = dict(row.source_binding_json or {}) or _automatic_binding(item, candidates)
            if binding is None:
                errors.append("SOURCE_CONFIRMATION_REQUIRED")
            else:
                try:
                    payload = _binding_payload(item, payload, binding)
                    row.source_binding_json = binding
                except QuestionImportError as exc:
                    errors.append(str(exc))
        row.payload_json = payload
        row.validation_errors_json = list(dict.fromkeys(errors))
        row.certification_report_json = {}
        if not errors:
            row.status = "pending"
            eligible.append((row.public_id, payload))
        elif "SOURCE_CONFIRMATION_REQUIRED" in errors:
            row.status = "source_confirmation_required"
        else:
            row.status = "template_invalid"
        if commit_progress:
            db.commit()

    def persist_batch(batch_results):
        nonlocal valid_count
        for row_id, result in batch_results.items():
            row = row_by_id[row_id]
            row.certification_report_json = result.report
            row.validation_errors_json = [] if result.issue_kind == "valid" else [result.issue_kind]
            row.status = result.issue_kind
            valid_count += int(result.issue_kind == "valid")
        run.valid_row_count = valid_count
        if commit_progress:
            db.commit()

    certify_question_payloads(eligible, on_batch_complete=persist_batch)
    run.valid_row_count = valid_count
    run.status = "ready_to_publish" if rows and valid_count == len(rows) else "needs_attention"
    run.error_summary = None if run.status == "ready_to_publish" else _attention_summary(rows)
    if commit_progress:
        db.commit()
    else:
        db.flush()
    return run


def queue_import_validation(db: Session, run: QuestionImportRun) -> QuestionImportRun:
    """Reset an unpublished run so a worker can persist row-level progress."""

    if run.status == "published":
        raise QuestionImportError("QUESTION_IMPORT_ALREADY_PUBLISHED")
    rows = list(db.scalars(select(QuestionImportRow).where(QuestionImportRow.run_id == run.id)))
    for row in rows:
        row.status = "pending"
        row.validation_errors_json = []
        row.certification_report_json = {}
    run.status = "validating"
    run.valid_row_count = 0
    run.error_summary = None
    db.commit()
    db.refresh(run)
    return run


def retry_question_import_rows(
    db: Session,
    run: QuestionImportRun,
    *,
    statuses: set[str],
    commit_progress: bool = True,
) -> QuestionImportRun:
    """Retry selected semantic outcomes without re-calling completed rows."""

    rows = list(
        db.scalars(select(QuestionImportRow).where(QuestionImportRow.run_id == run.id))
    )
    targets = [row for row in rows if row.status in statuses]
    if not targets:
        return run
    row_by_id = {row.public_id: row for row in targets}
    valid_count = sum(row.status in {"valid", "published"} for row in rows)
    run.status = "validating"
    run.valid_row_count = valid_count
    run.error_summary = None
    for row in targets:
        row.status = "pending"
        row.validation_errors_json = []
        row.certification_report_json = {}
    if commit_progress:
        db.commit()

    def persist_batch(batch_results):
        nonlocal valid_count
        for row_id, result in batch_results.items():
            row = row_by_id[row_id]
            row.certification_report_json = result.report
            row.validation_errors_json = [] if result.issue_kind == "valid" else [result.issue_kind]
            row.status = result.issue_kind
            valid_count += int(result.issue_kind == "valid")
        run.valid_row_count = valid_count
        if commit_progress:
            db.commit()

    certify_question_payloads(
        [(row.public_id, dict(row.payload_json or {})) for row in targets],
        on_batch_complete=persist_batch,
    )
    run.valid_row_count = sum(row.status in {"valid", "published"} for row in rows)
    run.status = (
        "ready_to_publish"
        if rows and run.valid_row_count == len(rows)
        else "needs_attention"
    )
    run.error_summary = None if run.status == "ready_to_publish" else _attention_summary(rows)
    if commit_progress:
        db.commit()
    else:
        db.flush()
    return run


def run_import_validation(run_id: str) -> None:
    try:
        with SessionLocal() as db:
            run = db.scalar(select(QuestionImportRun).where(QuestionImportRun.public_id == run_id))
            if run is None or run.status == "published":
                return
            validate_import_run(db, run, commit_progress=True)
    except Exception:
        logger.exception("question import validation failed run_id=%s", run_id)
        with SessionLocal() as db:
            run = db.scalar(select(QuestionImportRun).where(QuestionImportRun.public_id == run_id))
            if run is not None and run.status != "published":
                run.status = "needs_attention"
                run.error_summary = "题库认证任务异常，请重新校验"
                db.commit()
    finally:
        with _validation_lock:
            _active_validations.discard(run_id)


def run_import_row_retries(run_id: str, statuses: set[str]) -> None:
    try:
        with SessionLocal() as db:
            run = db.scalar(
                select(QuestionImportRun).where(QuestionImportRun.public_id == run_id)
            )
            if run is None or run.status == "published":
                return
            retry_question_import_rows(db, run, statuses=statuses)
    except Exception:
        logger.exception("question import row retry failed run_id=%s", run_id)
        with SessionLocal() as db:
            run = db.scalar(
                select(QuestionImportRun).where(QuestionImportRun.public_id == run_id)
            )
            if run is not None and run.status != "published":
                run.status = "needs_attention"
                run.error_summary = "题库认证任务异常，请重新校验"
                db.commit()
    finally:
        with _validation_lock:
            _active_validations.discard(run_id)


def schedule_import_validation(run_id: str) -> None:
    with _validation_lock:
        if run_id in _active_validations:
            return
        _active_validations.add(run_id)
    Thread(
        target=run_import_validation,
        args=(run_id,),
        name=f"question-import-validation-{run_id}",
        daemon=True,
    ).start()


def schedule_import_row_retries(run_id: str, statuses: set[str]) -> None:
    with _validation_lock:
        if run_id in _active_validations:
            return
        _active_validations.add(run_id)
    Thread(
        target=run_import_row_retries,
        args=(run_id, statuses),
        name=f"question-import-row-retry-{run_id}",
        daemon=True,
    ).start()


def recover_pending_import_validations() -> list[str]:
    with SessionLocal() as db:
        run_ids = list(
            db.scalars(
                select(QuestionImportRun.public_id).where(QuestionImportRun.status == "validating")
            )
        )
    for run_id in run_ids:
        schedule_import_validation(run_id)
    return run_ids


def set_row_source_binding(
    db: Session, *, run: QuestionImportRun, row_id: str, source_ref_ids: list[str], quotes: dict[str, str]
) -> QuestionImportRow:
    row = db.scalar(
        select(QuestionImportRow).where(QuestionImportRow.run_id == run.id, QuestionImportRow.public_id == row_id)
    )
    if row is None:
        raise QuestionImportError("QUESTION_IMPORT_ROW_NOT_FOUND")
    row.source_binding_json = {"source_ref_ids": source_ref_ids, "quotes": quotes}
    db.flush()
    return row


def publish_import_run(db: Session, run: QuestionImportRun) -> dict[str, Any]:
    if run.status == "published":
        return serialize_run(db, run)
    if run.status != "ready_to_publish":
        raise QuestionImportError("QUESTION_IMPORT_NOT_READY_TO_PUBLISH")
    items = {
        item.public_id: item
        for item in _catalog_items(db, run.domain_code, change_set_id=run.change_set_id)
    }
    rows = list(db.scalars(select(QuestionImportRow).where(QuestionImportRow.run_id == run.id)))
    for row in rows:
        existing = db.scalar(
            select(DiagnosticQuestion).where(
                DiagnosticQuestion.domain_code == run.domain_code,
                DiagnosticQuestion.external_id == row.question_external_id,
            )
        )
        item = items[row.knowledge_ref]
        payload = dict(row.payload_json or {})
        if existing is not None:
            same_content = (
                existing.knowledge_item_id == item.id
                and existing.question_type == payload["question_type"]
                and existing.stem == payload["stem"]
                and existing.options_json == payload["options"]
                and existing.difficulty == payload["difficulty"]
            )
            if not same_content:
                raise QuestionImportError("QUESTION_EXTERNAL_ID_ALREADY_EXISTS")
            row.status = "published"
            continue
        answer_key = {
            ("correct_option" if payload["question_type"] == "single_choice" else "answer"): payload["answer"],
            "explanation": payload["explanation"],
            "rubric": payload.get("rubric") or [],
            "question_slot": payload["question_slot"],
            "quiz_level": payload["quiz_level"],
            "question_bank_uses": payload["question_bank_uses"],
            "source_quote": payload["source_quote"],
            "source_ref_ids": payload["source_ref_ids"],
            "source_locators": payload["source_locators"],
            "source_content_hashes": payload["source_content_hashes"],
            "evidence_quotes": payload["evidence_quotes"],
            "chunker_version": payload["chunker_version"],
            "question_import_run_id": run.public_id,
            "question_import_row_id": row.public_id,
        }
        if run.change_set_id:
            answer_key["pending_change_set_id"] = run.change_set_id
        # A stale question is historical evidence, not a writable record.  A
        # replacement workbook therefore disables it only after the new,
        # independently certified question has been staged for publication.
        stale_slot_questions = list(
            db.scalars(
                select(DiagnosticQuestion).where(
                    DiagnosticQuestion.domain_code == run.domain_code,
                    DiagnosticQuestion.knowledge_item_id == item.id,
                    DiagnosticQuestion.status == "active",
                    DiagnosticQuestion.certification_status == "stale",
                )
            )
        )
        for stale_question in stale_slot_questions:
            stale_answer = dict(stale_question.answer_key_json or {})
            if str(stale_answer.get("question_slot") or "") != str(payload["question_slot"]):
                continue
            stale_question.status = "disabled"
            stale_question.disabled_at = datetime.now(UTC).replace(tzinfo=None)
            stale_question.disabled_reason = "已由认证题库替换"
        db.add(
            DiagnosticQuestion(
                public_id=f"dq_{uuid4().hex[:12]}",
                external_id=row.question_external_id,
                domain_code=run.domain_code,
                knowledge_item_id=item.id,
                related_knowledge_ids_json=[],
                question_type=payload["question_type"],
                stem=payload["stem"],
                options_json=payload["options"],
                answer_key_json=answer_key,
                difficulty=payload["difficulty"],
                status="staged" if run.change_set_id else "active",
                certification_status="certified",
                certification_rule_version=QUESTION_CERTIFICATION_RULE_VERSION,
                certification_report_json=row.certification_report_json,
                source_content_hash=payload["source_content_hash"],
                certified_at=datetime.now(UTC).replace(tzinfo=None),
            )
        )
        row.status = "published"
    run.status = "published"
    run.published_at = datetime.now(UTC)
    db.flush()
    domain = db.scalar(select(Domain).where(Domain.domain_code == run.domain_code))
    change_set = db.get(DomainChangeSet, run.change_set_id) if run.change_set_id else None
    if change_set is not None:
        remaining = question_gap_slots(
            db, run.domain_code, change_set_id=change_set.id
        )
        summary = dict(change_set.summary_json or {})
        summary["question_runs"] = list(
            dict.fromkeys([*(summary.get("question_runs") or []), run.public_id])
        )
        summary["remaining_question_slots"] = len(remaining)
        change_set.summary_json = summary
        change_set.status = "ready_to_activate" if not remaining else "questions_preparing"
        change_set.error_summary = None if not remaining else "题库仍有待补齐的缺口"
        readiness = {
            "passed": False,
            "status": "staged_for_change_set",
            "remaining_question_slots": len(remaining),
            "change_set_id": change_set.public_id,
        }
    else:
        readiness = DomainApiService(db).readiness(run.domain_code)
        if domain is not None:
            domain.status = "ready" if readiness["passed"] else "preparing"
    db.commit()
    return {**serialize_run(db, run), "readiness": readiness}


def serialize_row(row: QuestionImportRow) -> dict[str, Any]:
    payload = dict(row.payload_json or {})
    return {
        "row_id": row.public_id,
        "row_number": row.row_number,
        "question_external_id": row.question_external_id,
        "slot_key": row.slot_key,
        "knowledge_ref": row.knowledge_ref,
        "question_type": payload.get("question_type"),
        "difficulty": payload.get("difficulty"),
        "stem": payload.get("stem"),
        "options": list(payload.get("options") or []),
        "answer": payload.get("answer"),
        "explanation": payload.get("explanation"),
        "rubric": list(payload.get("rubric") or []),
        "purpose": next(iter(payload.get("question_bank_uses") or []), None),
        "quiz_level": payload.get("quiz_level"),
        "candidate_sources": row.candidate_sources_json or [],
        "source_binding": row.source_binding_json or {},
        "certification_report": row.certification_report_json or {},
        "status": row.status,
        "validation_errors": row.validation_errors_json or [],
        "issue_kind": None if row.status in {"valid", "published", "pending"} else row.status,
        "issue_fields": list((row.certification_report_json or {}).get("failed_fields") or []),
        "issue_reason": str((row.certification_report_json or {}).get("reason") or ""),
        "warnings": list((row.certification_report_json or {}).get("warnings") or []),
        "can_confirm_source": row.status == "source_confirmation_required",
    }


def serialize_run(db: Session, run: QuestionImportRun) -> dict[str, Any]:
    rows = list(db.scalars(select(QuestionImportRow).where(QuestionImportRow.run_id == run.id)))
    return {
        "run_id": run.public_id,
        "domain_code": run.domain_code,
        "template_version": run.template_version,
        "knowledge_catalog_fingerprint": run.knowledge_catalog_fingerprint,
        "question_inventory_fingerprint": run.question_inventory_fingerprint,
        "knowledge_scope": list(run.scope_json or []),
        "change_set_id": (
            db.scalar(select(DomainChangeSet.public_id).where(DomainChangeSet.id == run.change_set_id))
            if run.change_set_id else None
        ),
        "original_name": run.original_name,
        "status": run.status,
        "error_summary": run.error_summary,
        "row_count": run.row_count,
        "valid_row_count": run.valid_row_count,
        "processed_row_count": sum(row.status != "pending" for row in rows),
        "is_validating": run.status == "validating",
        "needs_attention_count": sum(
            row.status not in {"valid", "published", "pending"} for row in rows
        ),
        "source_confirmation_count": sum(
            row.status == "source_confirmation_required" for row in rows
        ),
        "content_rejected_count": sum(row.status == "content_rejected" for row in rows),
        "certification_service_error_count": sum(
            row.status == "certification_service_error" for row in rows
        ),
        "template_invalid_count": sum(row.status == "template_invalid" for row in rows),
        "published_at": run.published_at.isoformat() if run.published_at else None,
    }
